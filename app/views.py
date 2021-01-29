from flask_login import login_user, logout_user, current_user, login_required, LoginManager
#from app import app, db, lm, menu_views
from app import app, lm, db
#from app.menu_views import *
from flask import g,render_template, flash, redirect, session, Flask, url_for, request, jsonify
from .forms import LoginForm, AddProductForm, AmendProductForm, SearchForm, SellCash, MoveStock, AddCustomerForm, SellLoan, Sadad, CreateUser, EditVAT, Refund, Spendings,RevenueAccount, VATAccount,Procurement, MoveStockAdmin, MoveStockAdminShelf, Discount, Sadad2, SearchCustomer, EditUser, AssignShelf, AssignQuantityWarehouse, Sadad33
from .models import User, BranchOneProduct, BranchTwoProduct, CreditTransaction, DebitTransaction, Invoice, Product, Transaction, Inv, Account, VAT, Customer, RevenueTransaction, LoanTransaction
#from flask_table import Table, Col, LinkCol
from flask_wtf import Form as BaseForm
from functools import wraps
from passlib.hash import sha256_crypt
import smtplib
import os
import email.encoders
import email.mime.text
import email.mime.base
import base64
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from flask import jsonify
import re
import random
import string
from wtforms import StringField, PasswordField, FileField, BooleanField, TextAreaField, IntegerField, DateTimeField,SelectField,SelectMultipleField, DecimalField, validators
from wtforms.ext.sqlalchemy.fields import QuerySelectField
from wtforms.ext.sqlalchemy.fields import QuerySelectMultipleField
from wtforms import Form
from wtforms.validators import DataRequired, Email, EqualTo
from wtforms.fields.html5 import EmailField
from flask_sqlalchemy import SQLAlchemy
from functools import partial
from sqlalchemy import orm
from flask_wtf import FlaskForm
from flask import Flask, abort, request
import json
from werkzeug.datastructures import MultiDict
import time
from datetime import datetime
from passlib.hash import sha256_crypt
from sqlalchemy import exc
from datetime import datetime
#import xlrd
#import pandas as pd
import openpyxl 
from sqlalchemy.exc import IntegrityError
from sqlalchemy import desc
from datetime import date
from sqlalchemy.sql import func
from calendar import monthrange
#import hexdump
import numpy as np
#import datetime1 from 
#from urllib import urlencode, quote, unquote
#global id_number_for_form = 0;
ev_num = 0

items_list = []

def loadExcel():
  # Give the location of the file
  #loc = ("C:\\Users\\Reham\\Desktop\\aaaaaaaa.xlsx")
 
  #file = 'aaaaaaaa.xlsx'

  # Load spreadsheet
  #xl = pd.ExcelFile(file)

  # Print the sheet names
  #print(xl.sheet_names)

  # Load a sheet into a DataFrame by name: df1
  #df1 = xl.parse('Sheet1')

  # Give the location of the file 
  path = "C:\\Users\\Reham\\Desktop\\tabs4.xlsx"
    
  # To open the workbook  
  # workbook object is created 
  wb_obj = openpyxl.load_workbook(path)

    
  # Get workbook active sheet object 
  # from the active attribute 
  sheet_obj = wb_obj.active 
    
  # Cell objects also have a row, column,  
  # and coordinate attributes that provide 
  # location information for the cell. 
    
  # Note: The first row or  
  # column integer is 1, not 0. 
    
  # Cell object is created by using  
  # sheet object's cell() method. 
  m_row = sheet_obj.max_row
  print(m_row)
  for i in range(2, m_row + 1): 
    cell_obj = sheet_obj.cell(row = i, column = 1) # Product Name
    cell_obj1 =   sheet_obj.cell(row = i, column = 2) # Shelf
    cell_obj2 =   sheet_obj.cell(row = i, column = 5) # Single Price
    cell_obj3 =   sheet_obj.cell(row = i, column = 6) # Bulk Price
    cell_obj4 =   sheet_obj.cell(row = i, column = 7) # Product Expense

    print(str(cell_obj1.value))

    """

    print(cell_obj.value)
    print(str(cell_obj1.value))
    #print(hexdump.dump(cell_obj1.value), sep=":")
    #print(cell_obj1.value.hex())
    a = np.fromstring(cell_obj1.value, dtype=np.uint8)
    print(a)
    string = ""
    for character in cell_obj1.value:
      if character.isalnum():
        string = string + character
    print("After modifications: ")    
    print(string)
    #for character in cell_obj1.value:
      #print(character, character.encode('hex'))
    #print(cell_obj2.value)
    #print(cell_obj3.value)


    """

  # Print value of cell object  
  # using the value attribute 
    print("Excel Data")
    print(i)
    #print(cell_obj.value)
    product = Product(name=cell_obj.value, bulk_price = cell_obj3.value , bulk_bulk_price = 0, single_price=cell_obj2.value, single_expense = cell_obj4.value, bulk_bulk_expense = cell_obj4.value, bulk_expense= cell_obj4.value, shelf="", quantity=0)
    product1 = BranchOneProduct(name=cell_obj.value, bulk_price = cell_obj3.value , bulk_bulk_price = 0, single_price=cell_obj2.value, single_expense = cell_obj4.value, bulk_bulk_expense = cell_obj4.value, bulk_expense= cell_obj4.value, shelf=cell_obj1.value, quantity=0)
    product2 = BranchTwoProduct(name=cell_obj.value, bulk_price = cell_obj3.value , bulk_bulk_price = 0, single_price=cell_obj2.value, single_expense = cell_obj4.value, bulk_bulk_expense = cell_obj4.value, bulk_expense= cell_obj4.value, shelf="", quantity=0)
    #product2 = BranchOneProduct(name=cell_obj.value, bulk_price = cell_obj3.value , bulk_bulk_price = 0, single_price=cell_obj2.value, shelf=cell_obj1.value, quantity=0)
    ###product = Product.query.filter_by(name=cell_obj.value).first()
    ########product1 = BranchOneProduct.query.filter_by(name=cell_obj.value).first()
    ########product1.shelf = string
    #db.session.commit()
    ###product2 = BranchTwoProduct.query.filter_by(name=cell_obj.value).first()
    ###product.single_price = cell_obj2.value
    ###product.bulk_price = cell_obj3.value

    ###product1.single_price = cell_obj2.value
    ###product1.bulk_price = cell_obj3.value

    ###product2.single_price = cell_obj2.value
    ###product2.bulk_price = cell_obj3.value

    ###db.session.commit()


    db.session.add(product)
    db.session.add(product1)
    db.session.add(product2)
    try:
      db.session.commit()
    except IntegrityError as err:
      print("DUPLICATE VALUE CATCHED!")
      db.session.rollback()  

  return 0



def loadExcelCustomer():
  # Give the location of the file
  #loc = ("C:\\Users\\Reham\\Desktop\\aaaaaaaa.xlsx")
 
  #file = 'aaaaaaaa.xlsx'

  # Load spreadsheet
  #xl = pd.ExcelFile(file)

  # Print the sheet names
  #print(xl.sheet_names)

  # Load a sheet into a DataFrame by name: df1
  #df1 = xl.parse('Sheet1')

  # Give the location of the file 
  path = "C:\\Users\\Reham\\Desktop\\tab10.xlsx"
    
  # To open the workbook  
  # workbook object is created 
  wb_obj = openpyxl.load_workbook(path)

    
  # Get workbook active sheet object 
  # from the active attribute 
  sheet_obj = wb_obj.active 
    
  # Cell objects also have a row, column,  
  # and coordinate attributes that provide 
  # location information for the cell. 
    
  # Note: The first row or  
  # column integer is 1, not 0. 
    
  # Cell object is created by using  
  # sheet object's cell() method. 
  m_row = sheet_obj.max_row
  print(m_row)
  for i in range(2, m_row + 1): 
    cell_obj = sheet_obj.cell(row = i, column = 1) 
    cell_obj1 =   sheet_obj.cell(row = i, column = 2) 
  # Print value of cell object  
  # using the value attribute 
    print("Excel Data")
    print(i)
    print(cell_obj.value)
    #product = BranchTwoProduct(name=cell_obj.value, bulk_price = 0 , bulk_bulk_price = 0, single_price=0, shelf=cell_obj1.value, quantity=0)
    customer = Customer(name=cell_obj.value, mobile=0)
    db.session.add(customer)
    try:
      db.session.commit()
    except IntegrityError as err:
      print("DUPLICATE VALUE CATCHED!")
      db.session.rollback()  

  return 0

@app.route('/index.html', methods=['GET', 'POST'])
def dashboardRedirection():
  return redirect('dash')

@app.route('/index', methods=['GET', 'POST'])
def dashboard():
    form = LoginForm(request.form)
    print("Inside /login")
    print(form.errors)
    cust = Customer.query.filter_by(id=1).first()
    print("Customer remaning balance: ")
    #print(cust.remaining_balance)
    if form.validate():
        print("inside form validation")
        print(form.username.data)
        print(form.password.data)
        error = try_login(form.username.data, form.password.data)
        print(error)
        if error == True:
          return redirect('dash')
        else: 
          return "Unauthorized Access"


def zeroAccount():
  account = Account.query.filter_by(id=1).first()
  print("Current Balance: ")
  print(account.balance)
  account.balance = 0
  vat_account = Account.query.filter_by(id=2).first()
  vat_account.balance = 0
  revenue_account = Account.query.filter_by(id=3).first()
  revenue_account.balance = 0
  db.session.commit()

def deleteRows():
  db.session.query(CreditTransaction).delete()
  db.session.query(DebitTransaction).delete()
  db.session.query(Inv).delete()
  db.session.query(RevenueTransaction).delete()
  #db.session.query(Customer).delete()
  #db.session.query(Product).delete()
  #db.session.query(BranchOneProduct).delete()
  #db.session.query(BranchTwoProduct).delete()
  db.session.query(LoanTransaction).delete()
  db.session.commit()

@app.route('/dash', methods=['GET', 'POST'])
@login_required
def dash():
  rev = RevenueTransaction.query.all()
  print("Revenue Transactions: ")
  print(rev)
  cart =  {

}
  #session['cart'] = cart
  #db.session.drop_all()
  #db.session.commit()
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  account = Account.query.filter_by(id=1).first()
  big_half = account.balance * (75/100)
  small_half = account.balance * (25/100)
  dec = revenue(12)
  nov = revenue(11)
  october = revenue(10)
  sep = revenue(9)
  aug = revenue(8)
  jul = revenue(7)
  jun = revenue(6)
  may = revenue(5)
  april = revenue(4)
  march = revenue(3)
  feb = revenue(2)
  jan = revenue(1)

  print("January: ", jan)
  print("February:" , feb)
  print("March", march)
  print("April", april)
  print("May", may)
  print("June", jun)
  print("July", jul)
  print("August", aug)
  print("Septmber", sep)
  print("Ocober", october)
  print("November", nov)
  print("December", dec)

  q1 = jan + feb + march
  q2 = april + may + jun
  q3 = jul + aug + sep
  q4 = october + nov + dec

  return render_template('index.html', user = u, big_half = big_half,  small_half = small_half, jan = jan, feb = feb, march = march, april=april, may=may, jun = jun, jul=jul,
    aug=aug, sep = sep, october=october, nov=nov, dec=dec, q1=q1, q2=q2,q3=q3,q4=q4, username = u.username)

@app.route('/',  methods=['GET', 'POST'])
@app.route('/login', methods=['GET', 'POST'])
def login():
    #return render_template('login.html')
    #user = g.user
    db.create_all()
    #db.session.query(Customer).delete()
    #db.session.commit()
    ##prod = Product.query.filter_by(name="ادبتر متحرك BSP مركز سيلة 1/8*1/4").first()
    ##print("CHECKING PORD SHELF FORMAT: ")
    ##print(prod.shelf)
    ##a = np.fromstring(prod.shelf, dtype=np.uint8)
    ##print(a)
    
    ##prod.shelf = "as975"
    ##db.session.commit()
    ##print("CHECKING PORD SHELF FORMAT 2: ")
    ##print(prod.shelf)
    ##a = np.fromstring(prod.shelf, dtype=np.uint8)
    ##print(a)
    
    #db.session.query(BranchTwoProduct).delete()
    #db.session.commit()
    ##data = loadExcel()
    ###data = loadExcelCustomer()
    ##hashed_password = sha256_crypt.hash(str("123"))
    ##user = User(username="adminnn", hashed_password=hashed_password,admin = True, admin_alike=False, warehouse = False, branch1=False, branch2 = False, name = "Abdulrahman Sulimani",
    ##phone = "050" , branch = "Kilo 7 Branch")


    ##db.session.add(user)
    ##db.session.commit()

    ##account = Account(balance=0, description = "current Account")
    ##account2 = Account(balance=0, description = "vat Account")
    ##account3 = Account(balance=0, description = "Revenue Account")
    ##db.session.add(account)
    ##db.session.add(account2)
    ##db.session.add(account3)
    ##db.session.commit()
    ##vat_percentage = VAT(vat = 15)
    ##db.session.add(vat_percentage)
    ##db.session.commit()
    accounts = Account.query.all()
    print("Available accounts are: ")
    print(accounts)
    ###zeroAccount()
    ###deleteRows()
    form = LoginForm(request.form)
    print("Inside /login")
    account = Account.query.filter_by(id=1).first()
    print("Current Balance: &&&&&&&&&&&&&&&&&&&&&&&&&&&&&")
    print(account.balance)
    #user_ids = db.session.query(Product.name)
    #all_ids = user_ids.all()
    #print(all_ids)
    data = ""
    return render_template('login.html', form=form, data=data)



@app.route('/assign-shelf-2', methods=['GET', 'POST'])
def assignShelf():
  form = AssignShelf(request.form)
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  requested_branch = ""
  if form.submit.data:
      print("INSIDE THE NEWEST FUNCTION")
      print("Branch 2")
      product = BranchTwoProduct.query.filter_by(name = form.autocomp.data).first()
      product.shelf = form.shelf.data
      db.session.commit()
      flash(u'تمت المناقلة', 'success')
      return render_template('assign-shelf-2.html', form=form, user=u )
  return render_template('assign-shelf-2.html', form=form, user=u )


@app.route('/assign-quantity-warehouse', methods=['GET', 'POST'])
def assignQuantityWarehouse():
  form = AssignQuantityWarehouse(request.form)
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  requested_branch = ""
  if form.submit.data:
      print("INSIDE THE NEWEST FUNCTION")
      print("Branch 2")
      product = Product.query.filter_by(name = form.autocomp.data).first()
      product.quantity = form.quantity.data
      db.session.commit()
      flash(u'تمت المناقلة', 'success')
      return render_template('assign-quantity-warehouse.html', form=form, user=u )
  return render_template('assign-quantity-warehouse.html', form=form, user=u )




@app.route('/admin-product-movement', methods=['GET', 'POST'])
def adminProductMovement():
  form = MoveStockAdmin(request.form)
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  requested_branch = ""
  if form.submit.data:
    requested_branch = form.branch.data
    if requested_branch == "فرع كيلو 8":
      print("Branch 1")
      product = BranchOneProduct.query.filter_by(name = form.autocomp.data).first()
      product.quantity = form.quantity.data
      db.session.commit()
    elif requested_branch == "فرع الجوهرة":
      print("Branch 2")
      product = BranchTwoProduct.query.filter_by(name = form.autocomp.data).first()
      product.quantity = form.quantity.data
      db.session.commit()
    elif requested_branch == "المستودع":
      print("warehouse")
      product = Product.query.filter_by(name = form.autocomp.data).first()
      product.quantity = form.quantity.data
      db.session.commit()

    flash(u'تمت المناقلة', 'success')
    return render_template('admin-product-movement.html', form=form, user=u )
  return render_template('admin-product-movement.html', form=form, user=u )


@app.route('/admin-product-movement-shelf', methods=['GET', 'POST'])
def adminProductMovementShelf():
  form = MoveStockAdminShelf(request.form)
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  requested_branch = ""
  if form.check.data:
    print("Sanity Check ...")
    print(form.autocomp.data)
    #product_test = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
    #print("Product Test: ")
    #print(product_test)
    #print(product_test.shelf)
    print("******************************************************************")
    print("******************************************************************")
    #product = BranchOneProduct.query.filter_by(shelf=form.autocomp.data).all()
    #product2 = BranchOneProduct.query.filter_by(shelf=form.autocomp.data).all()
    #product3 = BranchTwoProduct.query.filter_by(shelf=form.autocomp.data).all()
    #print(product)
    #print(product2)
    #print(product3)

    product=None
    requested_branch = form.branch.data
    if requested_branch == "فرع كيلو 8":
      print("Branch 1")
      product = BranchOneProduct.query.filter_by(shelf=form.autocomp.data).all()
    elif requested_branch == "فرع الجوهرة":
      print("Branch 2")
      product = BranchTwoProduct.query.filter_by(shelf=form.autocomp.data).all()

    prod_list=[]
    for prod in product:
      prod_list.append(prod.name)
      prod_list.append(prod.name)


    """  
    it = iter(prod_list)
    res_dct = list(zip(it, it))
    print("RRRREERRRREEEE TAAAATTTTOOOO")
    print(res_dct)
    print(type(res_dct))
    #list33 = [(k, v) for k, v in res_dct.items()] 
    listt = zip(res_dct.keys(), res_dct.values()) 
    print("MAAAAAA LIST 1111")
    print(listt)
    mmm = tuple(listt)
    print("MAAA TUPLE IZZZZ:  ")
    print(mmm)
    listt = list(mmm)
    print("MAAAAAA LIST 2222")
    print(listt)
    """
    mylist=[]
    res = tuple(map(tuple, prod_list))
    ress = tuple(prod_list)
    print("CHEEECKING .... ")
    print(ress)
    mylist.append(ress)
    print(mylist)
    form.products.choices =  mylist
    return render_template('admin-product-movement-shelf.html', form=form, user=u )
  if form.submit.data:
    requested_branch = form.branch.data
    if requested_branch == "فرع كيلو 8":
      print("Branch 1")
      product = BranchOneProduct.query.filter_by(name = form.products.data).first()
      product.quantity = form.quantity.data
      db.session.commit()
    elif requested_branch == "فرع الجوهرة":
      print("Branch 2")
      product = BranchTwoProduct.query.filter_by(name = form.products.data).first()
      product.quantity = form.quantity.data
      db.session.commit()

    flash(u'تمت المناقلة', 'success')
    return render_template('admin-product-movement-shelf.html', form=form, user=u )
  return render_template('admin-product-movement-shelf.html', form=form, user=u )



@app.route('/customer-statement', methods=['GET', 'POST'])
def customerStatement():
  print("Inside Customer's statement")
  form = SearchCustomer(request.form)
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  if form.confirm.data:
    cust_name = form.autocompcustomer.data
    print("Customer name: ")
    print(cust_name)
    customer = Customer.query.filter_by(name=cust_name).first()
    print("Customer Object: ")
    print(customer)
    loan_transactions = LoanTransaction.query.filter_by(customer_id=customer.id).all()
    print("Loan Transactions: ")
    print(loan_transactions)
    return render_template('customer-statement.html', form=form, user=u, loan_transactions = loan_transactions, len = len(loan_transactions)) 
  customer = Customer.query.filter_by(name="Wessam Gholam").first()
  list_of_customers=[]
  return render_template('customer-statement.html', form=form, user=u, list_of_customers = list_of_customers, list_of_balances = list_of_customers, len = len(list_of_customers)) 



@app.route('/customer-statement1111', methods=['GET', 'POST'])
def customerStatement11():
  print("Inside Customer's statement")
  form = Procurement(request.form)
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  customer = Customer.query.filter_by(name="Wessam Gholam").first()

  list_of_customers = []
  list_of_balances = []
  customers = Customer.query.all()
  #invs = Inv.query.filter_by(customer_id = customer.id, remaining_balance >= form.pay_amount.data).all()
  for customer in customers: 
    invs = Inv.query.filter(Inv.remaining_balance > 0, Inv.inv_type == "Loan", Inv.customer_id == customer.id).all()
    total_balance = 0
    for inv in invs:
      total_balance = total_balance + inv.remaining_balance
    if total_balance > 0:
      list_of_customers.append(customer)
      list_of_balances.append(total_balance)

  print("Customer")
  print(customer)
  invoices = Inv.query.filter_by(customer_id= customer.id).all()
  print("Invoice")
  print(invoices)
  for inv in invoices:
    print("Credit Transaction: ")
    crs = CreditTransaction.query.filter_by(invoice_id = inv.id).all()
    print(crs)
  return render_template('customer-stmt-list.html', form=form, user=u, list_of_customers = list_of_customers, list_of_balances = list_of_balances, len = len(list_of_balances)) 



@app.route('/procurement', methods=['GET', 'POST'])
def procurement():
  form = Procurement(request.form)
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  if form.submit.data:
    print("SUBMIT")
    print("My invoice selections are: ")
    print(form.invoices_to_choose.data)
    customer = Customer.query.filter_by(name=form.autocompcustomer.data).first()
    print("Customer")
    print(customer)
    invs = customer.invoices
    rem_balance = 0
    print("Pay amount")
    print(form.pay_amount.data)
    print(form.invoices_to_choose.data.split('--- ')[1]) 
    wanted_inv_id = int(form.invoices_to_choose.data.split('--- ')[1])
    wanted_inv = Inv.query.filter_by(id=wanted_inv_id).first()
    wanted_inv.remaining_balance = wanted_inv.remaining_balance - float(form.pay_amount.data)
    db.session.commit()
    flash(u'تمت العملية بنجاح', 'success')
    return render_template('procurement.html', form=form, user=u)  

  if form.invoices.data:
    customer = Customer.query.filter_by(name=form.autocompcustomer.data).first()
    #invs = Inv.query.filter_by(customer_id = customer.id, remaining_balance >= form.pay_amount.data).all()
    invs = Inv.query.filter(Inv.remaining_balance >= form.pay_amount.data, Inv.customer_id == customer.id).all()
    numb = 1
    print("My dicts")
    print(form.invoices_to_choose.choices)
    for inv in invs:
      #form.invoices_to_choose.choices['numb'] = "فاتورة: " + str(inv.id) + " --- " + "رصيد متبقي" + " "+ str( inv.remaining_balance)
      #form.invoices_to_choose.choices.append("فاتورة: " + str(inv.id) + " --- " + "رصيد متبقي" + " "+ str( inv.remaining_balance))
      form.invoices_to_choose.choices.append("SAR " + str(inv.remaining_balance) +  " --- " + str(inv.id))

      numb = numb + 1
    print(invs)
    return render_template('procurement.html', form=form, user=u)  

  return render_template('procurement.html', form=form, user=u)

@app.route('/users', methods=['GET', 'POST'])
def users():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  form = CreateUser(request.form)
  if form.submit.data:
    print("USER SUBMITTED.. ")
    print(form.admin.data)
    print(form.admin_alike.data)
    print(form.warehouse.data)
    print(form.branch1.data)
    print(form.branch2.data)
    hashed_password = sha256_crypt.hash(str(form.password.data))
    user = User(username=form.username.data, hashed_password=hashed_password,admin = form.admin.data, admin_alike = form.admin_alike.data, 
    warehouse = form.warehouse.data, branch1 = form.branch1.data, branch2= form.branch2.data, name = "....",
    phone = "050" , branch = "Kilo 7 Branch")
    db.session.add(user)
    db.session.commit()
    users = User.query.all()
    flash(u'تمت اضافة المستخدم', 'success')
    return render_template('users.html', users=users, len=len(users), form=form, user=u)
  users = User.query.all()
  return render_template('users.html', users=users, len=len(users), form=form, user=u)

@app.route('/delete-users/<user_id>', methods=['GET', 'POST'])
def deleteUsers(user_id):
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  user = User.query.filter_by(id=user_id).first()
  db.session.delete(user)
  db.session.commit()
  users = User.query.all()
  flash(u'تمت اضافة المستخدم', 'success')
  return redirect(url_for('users'))


@app.route('/edit-users/<user_id>', methods=['GET', 'POST'])
def editUsers(user_id):
  form = EditUser(request.form)
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  user = User.query.filter_by(id=user_id).first()
  form.username1.data = user.username
  if form.submit.data:
    print("Inside submit clause: ")
    print(form.username1.data)
    print(form.password1.data)
    user = User.query.filter_by(id=user_id).first()
    print("Amended user: ")
    print(user)
    print(user.username)
    print("Data from form: ")
    print(form.username1.data)
    password = form.password1.data
    hashed_password = sha256_crypt.hash(str(password))
    user.hashed_password = hashed_password
    db.session.commit()
    print("After amendment: ")
    print(user.username)
    return redirect(url_for('users'))
  return render_template('edit-user.html', myuser=user.id,form=form, user=u)


#@app.route('/',  methods=['GET', 'POST'])
#@app.route('/login', methods=['GET', 'POST'])
#def login():
#    return render_template('login.html')
 #   user = g.user
  #  form = LoginForm()
   # if form.validate_on_submit():
    #    error = try_login(form.username.data, form.password.data)
     #   if not error:
      #      session['logged_in'] = True
       #     if user.admin:
        #        session['admin'] = True
         #   return redirect('/index')
    #return render_template('login.html', form=form)


# Logout
@app.route("/logout")
@login_required
def logout():
    logout_user()
    session['logged_in'] = False
    session['admin'] = False
    return redirect('/login')

NAMES=["abc","abcd","abcde","abcdef"]


@app.route('/autocompleteshelfs',methods=['GET'])
def autocompleteshelfs():
    srch =[]
    print("####################################################################################")
    print("####################################################################################")
    print("####################################################################################")
    print("Inside autocomplete Shelfs")
    search = request.args.get('autocomplete')
    #print("Search words")
    #print(search)
    #print(NAMES)
    product_list = []
    products = Product.query.all()
    for p in products:
      if p.shelf:
        #print(p.shelf)
        product_list.append(p.shelf)
    #print(product_list)  
    #query = request.args.get('query')
    app.logger.debug(search)
    return jsonify(json_list=product_list)


@app.route('/autocomplete',methods=['GET'])
def autocomplete():
    srch =[]
    print("Inside autocomplete")
    search = request.args.get('autocomplete')
    #print("Search words")
    #print(search)
    #print(NAMES)
    product_list = []
    products = Product.query.all()
    for p in products:
      product_list.append(p.name)
    #print(product_list)  
    #query = request.args.get('query')
    app.logger.debug(search)
    return jsonify(json_list=product_list)


@app.route('/autocompleteb1',methods=['GET'])
def autocompleteb1():
    srch =[]
    print("Inside autocomplete")
    search = request.args.get('autocomplete')
    #print("Search words")
    #print(search)
    #print(NAMES)
    product_list = []
    products = BranchOneProduct.query.all()
    for p in products:
      product_list.append(p.name)
    #print(product_list)  
    #query = request.args.get('query')
    app.logger.debug(search)
    return jsonify(json_list=product_list)    

@app.route('/autocomplete-customer',methods=['GET'])
def autocompletecustomer():
    srch =[]
    print("Inside autocomplete")
    search = request.args.get('autocomplete')
    #print("Search words")
    #print(search)
    #print(NAMES)
    customer_list = []
    customers = Customer.query.all()
    for c in customers:
      customer_list.append(c.name)
    #print(customer_list)  
    #query = request.args.get('query')
    app.logger.debug(search)
    return jsonify(json_list=customer_list)



# @app.route('/autocomplete',methods=['GET'])
# def autocomplete():
#     srch =[]
#     print("Inside autocomplete")
#     search = request.args.get('autocomplete')
#     print("Search words")
#     print(search)
#     print(NAMES)
#     product_list = []
#     products = Product.query.all()
#     for p in products:
#       product_list.append(p.name)
#     print(product_list)  
#     query = request.args.get('query')
#     inn = 1
#     #search = request.args.get('autocomplete')
#     if search is not None:
#         # do some stuff to open your names text file
#         # do some other stuff to filter
#         # put suggestions in this format...
        
#         suggestions_list = ["joe","jim","fibi","wiz","wessam", "وسام", "وسيم"]
#         user_ids = db.session.query(Product.name)
#         suggestions = user_ids.all()
#         print(suggestions_list)
#         print(suggestions_list[0])
#         print(type(suggestions_list))
#         print(suggestions)
#         print(suggestions[0][0])
#         print(type(suggestions))
#         index=0
#         for sug in product_list:
#           print("inside for loop")
#           print(sug)
#           print(type(sug))
#           res = sug.startswith( search, 0, len(sug) )
#           print("before if")
#           if res == True:
#             print("inside if")
#             srch.append(sug)
#             #index = index+1
#             print(srch)

#     app.logger.debug(search)
#     return jsonify(json_list=srch) 


@app.route("/current-balance")
@login_required
def currentBalance():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  form = RevenueAccount()
  account = Account.query.filter_by(id=1).first()
  print("Current Balance BEFORE refund @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@: ")
  print(account.balance)
  form.balance.data = account.balance
  return render_template('current-balance.html', form=form, user=u)


@app.route("/current-revenue")
@login_required
def currentRevenueBalance():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  form = RevenueAccount()
  account = Account.query.filter_by(id=3).first()
  print("Current Revenue Balance @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@: ")
  print("::::::::::::::::::::::::................:::::::::::::........:::::::::::")
  print(account.balance)
  form.balance.data = account.balance
  return render_template('current-revenue.html', form=form, user=u) 

@app.route("/vat-balance",methods=['GET','POST'])
@login_required
def VATBalance():
  print("Inside VAT")
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  form = VATAccount(request.form)
  account = Account.query.filter_by(id=2).first()
  print("Current Balance BEFORE refund @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@: ")
  print(account.balance)
  form.balance.data = account.balance
  if form.submit.data:
    print("At submit")
    now = datetime.now()
    amount = form.vat_amount.data
    account.balance = account.balance - float(amount)
    db.session.commit()
    dr = DebitTransaction(t_type = "DR", total=amount, date=now, description="تسديد الضريبة", invoice_id=None, current_balance= account.balance)
    db.session.add(dr)
    db.session.commit()
    flash(u'تم التسديد بنجاح', 'success')
  return render_template('vat-account.html', form=form,user=u) 

@app.route("/statement")
@login_required
def statement():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  products = []
  now = datetime.now()
  print("Date & Time now")
  print(now)
  #cr = CreditTransaction(t_type="CR", total=44, date=now, description="selling")
  #dr = DebitTransaction(t_type="DR", total=10, date=now, description="selling")

  #db.session.add(cr)
  #db.session.add(dr)
  #db.session.commit()
  #DebitTransaction.query.delete()
  #db.session.commit()
  crs = CreditTransaction.query.all()
  drs = DebitTransaction.query.all()

  print("CRs")
  print(crs)
  print("DRs")
  print(drs)
  crs.extend(drs)
  print("CR & DR together")
  print(crs) 
  return render_template('statement.html', products=crs, len=len(crs),user=u)

@app.route("/search/<string:box>")
def process(box):
    #user_ids = session.query(Product.name)
    #all_ids = user_ids.all()
    #print(all_ids)
    product_list = []
    products = Product.query.all()
    for p in products:
      product_list.append(p.name)
    print(product_list)  
    query = request.args.get('query')
    search = request.args.get('autocomplete')
    if box == 'names':
        # do some stuff to open your names text file
        # do some other stuff to filter
        # put suggestions in this format...
        srch =[]
        suggestions_list = ["joe","jim","fibi","wiz","wessam", "وسام", "وسيم"]
        user_ids = db.session.query(Product.name)
        suggestions = user_ids.all()
        print(suggestions_list)
        print(suggestions_list[0])
        print(type(suggestions_list))
        print(suggestions)
        print(suggestions[0][0])
        print(type(suggestions))
        index=0
        for sug in product_list:
          print("inside for loop")
          print(sug)
          print(type(sug))
          res = sug.startswith( query, 0, len(sug) )
          print("before if")
          if res == True:
            print("inside if")
            srch.append(sug)
            #index = index+1
            print(srch)
    if box == 'songs':
        # do some stuff to open your songs text file
        # do some other stuff to filter
        # put suggestions in this format...
        suggestions = [{'value': 'song1','data': '123'}, {'value': 'song2','data': '234'}]
    return Response(json.dumps(srch), mimetype='application/json')
    #return jsonify({"suggestions":srch})  

@app.route("/search")
@login_required
def search():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  form = SearchForm(request.form)
  return render_template('search.html', form=form, user=u)
  

@app.route("/move-stock.html")
@login_required
def moveStock():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  form = MoveStock(request.form)
  return render_template('move-stock.html', form=form, user=u)

@app.route("/fetch-stock.html", methods=['GET', 'POST'])
@login_required
def fetchStock():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  form = MoveStock(request.form)
  print("Check")
  print(form.autocomp.data)
  print(form.check.data)
  print(form.submit.data)
  prod = form.autocomp.data
  p = Product.query.filter_by(name=form.autocomp.data).first()
  if form.check.data == True:
    print("Inside Check button")
    print(p.quantity)
    form.checking.data = p.quantity
    form.autocomp.data = form.autocomp.data
    return render_template('move-stock.html', form=form, user=my_user)
  if form.submit.data == True:
    print("Inside Submit button")
    print(form.number.data) 
    print(form.branch.data)
    if p.quantity == 0:
      flash(u'Product Quantity is out of stock', 'danger')
      return redirect('move-stock.html')
    if form.branch.data == "Branch 1": 
      print("insode branch 1")
      b1 = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
      b1.quantity = b1.quantity + int(form.number.data)
      p.quantity = p.quantity - int(form.number.data)
    else:
      print("insode branch 2")
      b2 = BranchTwoProduct.query.filter_by(name=form.autocomp.data).first()
      b2.quantity = b2.quantity + int(form.number.data)
      p.quantity = p.quantity - int(form.number.data)
    db.session.commit()
  #if form.validate():
   # print("inside form validation")
    #print(form.autocomp.data)
    #print(form.number.data)
    #print(form.check.data)
    #print(form.submit.data)
  flash(u'تمت المناقلة بنجاح', 'success')
  form = MoveStock(request.form)
  return render_template('move-stock.html', form=form, user=u)

@app.route("/invoice/edit/<invoice_id>", methods=['GET', 'POST'])
@login_required
def invoiceEdit(invoice_id):
  form = Discount(request.form)
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  p = Inv.query.filter_by(id=invoice_id).first()
  print("Invoice Data: ")
  print(p.products);
  print(type(p.products))
  print(p.inv_type)
  print(p.status)
  print(p.remaining_balance)
  curr_customer = None
  if p.customer_id != None:
    print(p.customer_id)
    curr_customer = Customer.query.filter_by(id=p.customer_id).first().name
    print(curr_customer)
  else:
    print("customer_id is None")  
  list1 = json.loads(p.products) 
  list_of_items = []
  total = 0

  if form.submit.data:

    print("INSIDE DISCOUNT MODE")
    print(float(form.discount_amount.data))
    print(p.total)
    new_total = p.total - float(form.discount_amount.data)
    vat_value = VAT.query.all()
    vat_value = vat_value[0].vat
    vat = vat_value * new_total / 100
    print("VAT CALCULATED: ")
    print(vat)
    new_total = new_total - vat
    print(new_total)
    new_total = round(new_total, 2)
    p.total = new_total
    p.vat_value = vat 
    db.session.commit()
    print(p.total)
    #reCalculateInvoice(invoice_id)

    list1 = json.loads(p.products) 
    list_of_items = []
    total = 0
    for key, value in list1.items():
      temp = [key,value]
      list_of_items.append(temp)
    return render_template('edit-invoice.html', form=form, products=list_of_items, length= len(list_of_items), total = p.total, vat=p.vat_value, vat_percentage = p.vat_percentage, customer=curr_customer, remaining_balance= p.remaining_balance, category = p.category, user=u, invoice_id= invoice_id)

  for key, value in list1.items():
    temp = [key,value]
    list_of_items.append(temp)
  #return invoice_id;
  return render_template('edit-invoice.html', form=form, products=list_of_items, length= len(list_of_items), total = p.total, vat=p.vat_value, vat_percentage = p.vat_percentage, customer=curr_customer, remaining_balance= p.remaining_balance, category = p.category, user=u, invoice_id= invoice_id)
  #return render_template('edit-invoice.html',invoice_id=invoice_id)






@app.route("/invoice/refund/<invoice_id>", methods=['GET', 'POST'])
@login_required
def invoiceRefund(invoice_id):
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  debit_amount = 0
  print("Inside refund func()")
  form = Refund(request.form)
  p = Inv.query.filter_by(id=invoice_id).first()
  print("CHECK BOX VALUES ... ")
  print(form.refund_type.data)
  if form.submit.data and form.refund_type.data == None:
    print("TYPE IS NONE!!")
    flash(u'الرجاء اختيار نوع الاسترجاع', 'danger')
    return redirect(url_for('invoiceRefund', invoice_id=p.id)) 
  if p.inv_type == "Cash" or p.inv_type == "كبس"  or p.inv_type == "شبكة":
    if form.refund_type.data == "Full":
      account = Account.query.filter_by(id=1).first()
      vat_account = Account.query.filter_by(id=2).first()
      print("Current Balance BEFORE refund @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@: ")
      print(account.balance)
      print("Refund amount: ")
      print(form.refund_amount.data)
      total_after_vat = p.total - p.vat_value
      account.balance = account.balance - total_after_vat
      vat_account.balance = vat_account.balance - p.vat_value
      db.session.commit()
      print("Current Balance AFTER refund @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@: ")
      print(account.balance)
      now = datetime.now()
      dr = DebitTransaction(t_type="DR", total=total_after_vat, date=now, description="استرجاع قيمة فاتورة رقم  "+ str(p.id) , invoice_id=invoice_id, current_balance = account.balance)
      db.session.add(dr)
      p.is_refunded = True
      list1 = p.products
      print("Purchased products: ")
      print(p.products)
      print(type(p.products))
      y = json.loads(p.products) 
      print(y)     
      list_of_items = []
      for key, value in y.items():
        #temp = [key]
        temp = [key,value]
        list_of_items.append(temp)
        if p.initiator == "Branch1":
          product = BranchOneProduct.query.filter_by(name = key).first()
          product.quantity = product.quantity + int(value[1])
          db.session.commit()
        elif p.initiator == "Branch2":
          product = BranchTwoProduct.query.filter_by(name = key).first()
          product.quantity = p.quantity + int(value[1])  


      
      paid = 0
      revenue = 0
      revenue_total = 0       
      for item in list_of_items: 
        print(item)
        print(item[0])
        print(item[1])
        print(item[1][0])
        print(item[1][1])

        product = Product.query.filter_by(name=item[0]).first()
        paid = item[1][0] * item[1][1]
        if p.category == "تجزئة":
          print("Single")
          revenue = product.single_expense
        elif p.category == "جملة الجملة":
          print("Bulk Bulk")
          revenue = product.bulk_bulk_expense
        elif p.category == "جملة":
          print("Bulk")
          revenue = product.bulk_expense
        else:
          print("INOVICE TYPE IS NON OF THE ABOVE")  

        print(paid)
        revenue = revenue * item[1][1]
        print(revenue)
        revenue = paid - revenue
        print(revenue)
        revenue_total += revenue
        print(paid)
        print(revenue)
        print(revenue_total)
      revenue_account = Account.query.filter_by(id=3).first()
      revenue_account.balance = revenue_account.balance - revenue_total  
      print("Purchased products as a list: ")
      print(list_of_items)
      db.session.commit()
      flash(u'تم اعادة مبلغ الفاتورة', 'success')
      return redirect(url_for('viewInvoices'))
    elif form.refund_type.data == "Partial":
      account = Account.query.filter_by(id=1).first()
      vat_account = Account.query.filter_by(id=2).first()
      print("Current Balance: ")
      print(account.balance)
      print("Refund amount: ")
      print(form.refund_amount.data)
      list1 = p.products
      print("Purchased products: ")
      print(p.products)
      print(type(p.products))
      y = json.loads(p.products) 
      print(y)     
      list_of_items = []
      for key, value in y.items():
        #temp = [key]
        temp = [key,value]
        list_of_items.append(temp)

      total_refund_amount = 0  
      if p.initiator == "Branch1":
        print("Items")
        print(y)
        print(type(y))
        prod = y[form.refund_products.data]
        print("prod")
        print(prod)
        print(prod[0])
        qunatity = prod[1]
        price = prod[0]
        print(qunatity)
        print(price)
        total_refund_amount = float(form.refund_amount.data) * price
        #vat_of_refund_amount = (total_refund_amount / p.total)
        #vat_of_refund_amount = 100 * (p.vat_percentage / total_refund_amount)
        vat_of_refund_amount = (p.vat_percentage * total_refund_amount) / 100
        print("Total of vat refund amount 1: ")
        print(vat_of_refund_amount)
        #vat_of_refund_amount = vat_of_refund_amount * p.vat_value
        #print("Total of vat refund amount 2: ")
        #print(vat_of_refund_amount)
        print("Total Refund Amount : ")
        print(total_refund_amount)
        #print("Total of vat refund amount: ")
        #print(vat_of_refund_amount)
        product = BranchOneProduct.query.filter_by(name = key).first()
        product.quantity = product.quantity + int(form.refund_amount.data)
        db.session.commit()
        #return 'Ok'
      elif p.initiator == "Branch2":
        prod = list_of_items[form.refund_products.data]
        qunatity = prod[1][1]
        price = prod[1][0] 
        print(qunatity)
        print(price)
        total_refund_amount = qunatity * price
        print("Total Refund Amount : ")
        print(total_refund_amount)
        product = BranchTwoProduct.query.filter_by(name = key).first()
        product.quantity = p.quantity + qunatity
        db.session.commit()  
        return 'ok' 
      account.balance = account.balance - total_refund_amount
      vat_account.balance = vat_account.balance - vat_of_refund_amount

      paid = 0
      revenue = 0
      revenue_total = 0       
      qunatity = prod[1]
      price = prod[0]
      product = Product.query.filter_by(name=item[0]).first()
      paid = qunatity * price
      if p.category == "تجزئة":
        print("Single")
        revenue = product.single_expense
      elif p.category == "جملة الجملة":
        print("Bulk Bulk")
        revenue = product.bulk_bulk_expense
      elif p.category == "جملة":
        print("Bulk")
        revenue = product.bulk_expense
      else:
        print("INOVICE TYPE IS NON OF THE ABOVE")  

      print(paid)
      revenue = revenue * qunatity
      print(revenue)
      revenue = paid - revenue
      print(revenue)
      revenue_total += revenue
      print(paid)
      print(revenue)
      print(revenue_total)
      revenue_account = Account.query.filter_by(id=3).first()
      revenue_account.balance = revenue_account.balance - revenue_total  

      db.session.commit()
      now = datetime.now()
      dr = DebitTransaction(t_type="DR", total=total_refund_amount, date=now, description="استرجاع قيمة فاتورة رقم  "+ str(p.id) , invoice_id=invoice_id, current_balance = account.balance)
      db.session.add(dr)
      p.is_refunded = True
      db.session.commit()
      flash(u'تم اعادة مبلغ الفاتورة', 'success')
      return redirect(url_for('invoiceRefund', invoice_id=p.id)) 
      #return redirect(url_for('viewInvoices'))

  if p.inv_type == "Loan":
    debit_amount = 0
    if form.refund_type.data == "Full":
      debit_amount = 0
      debit_amount = p.total - p.remaining_balance
      print("DEBIT AMOUNT: ")
      print(debit_amount)
      vat_of_refund_amount = (p.vat_percentage * debit_amount) / 100
      print("VAT AMOUNT: ")
      print(vat_of_refund_amount)
      debit_amount = debit_amount - vat_of_refund_amount
      print("DEBIT AMOUNT 2: ")
      print(debit_amount)
      p.remianing_balance = 0
      account = Account.query.filter_by(id=1).first()
      vat_account = Account.query.filter_by(id=2).first()
      print("Current Balance: ")
      print(account.balance)
      print("Refund amount: ")
      print(form.refund_amount.data)
      account.balance = account.balance - debit_amount
      vat_account.balance = vat_account.balance - vat_of_refund_amount

      #sadad_percentage = float(p.remaining_balance) / float(p.total)
      #print("SADAD Percentage")
      #print(sadad_percentage)
      #vat = p.vat_value * (sadad_percentage)
      #print("VAT")
      #print(vat)
      #print("VAT AMOUNT: ")
      #print(vat_of_refund_amount)
      #sadad_amount = float(form.pay_amount.data) - vat



      y = json.loads(p.products) 
      print(y)     
      list_of_items = []
      for key, value in y.items():
        #temp = [key]
        temp = [key,value]
        list_of_items.append(temp)
        if p.initiator == "Branch1":
          product = BranchOneProduct.query.filter_by(name = key).first()
          product.quantity = product.quantity + int(value[1])
          db.session.commit()
        elif p.initiator == "Branch2":
          product = BranchTwoProduct.query.filter_by(name = key).first()
          product.quantity = p.quantity + int(value[1])
      db.session.commit()
      now = datetime.now()
      dr = DebitTransaction(t_type="DR", total=debit_amount, date=now, description="استرجاع قيمة فاتورة رقم  "+ str(p.id) , invoice_id=invoice_id, current_balance = account.balance)
      db.session.add(dr)
      p.is_refunded = True
      db.session.commit()
      flash(u'تم اعادة مبلغ الفاتورة', 'success')
      return redirect(url_for('invoiceRefund', invoice_id=p.id)) 
    elif form.refund_type.data == "Partial":
      y = json.loads(p.products) 
      print(y)   
      list_of_items = []
      for key, value in y.items():
        #temp = [key]
        temp = [key,value]
        list_of_items.append(temp)
        quantity = 0
        price = 0
        total_refund_amount = 0  
        if p.initiator == "Branch1":
          prod = y[form.refund_products.data]
          qunatity = prod[1]
          price = prod[0]
          product = BranchOneProduct.query.filter_by(name = key).first()
          product.quantity = product.quantity + int(form.refund_amount.data)

        if p.initiator == "Branch2":
          prod = y[form.refund_products.data]
          qunatity = prod[1]
          price = prod[0]
          product = BranchTwoProduct.query.filter_by(name = key).first()
          product.quantity = product.quantity + int(form.refund_amount.data)


      print("Price")
      print(price)

      amount = price * float(form.refund_amount.data) # Quantity entered
      print("Amount")
      print(amount)
      print("Remaining Balance")
      print(p.remaining_balance)
      print("Comparison caluse")
      print(amount <= p.remaining_balance)
      if amount <= p.remaining_balance:
        p.remaining_balance = float(p.remaining_balance) - float(amount)
      elif amount > p.remaining_balance:
        print("Breaking Bad ...")    
        print("amount to be refunded: ")
        print(amount)
        amount = amount - p.remaining_balance
        vat_of_refund = amount / p.total
        vat_of_refund = vat_of_refund * p.vat_value
        amount = amount -  vat_of_refund
        debit_amount = amount
        print("Amount to be debited from curr account: ")
        print(debit_amount)
        print("Amount to be debited from vat account: ")
        print(vat_of_refund)
        p.remaining_balance = 0
        account = Account.query.filter_by(id=1).first()
        vat_account = Account.query.filter_by(id=2).first()
        print("Current Balance: ")
        print(account.balance)
        print("Refund amount: ")
        print(form.refund_amount.data)
        account.balance = account.balance - debit_amount
        vat_account.balance = vat_account.balance - vat_of_refund
        db.session.commit()
        print("account balance")
        print(account.balance)
        print("vat_account balance")
        print(vat_account.balance)
        now = datetime.now()
        if debit_amount > 0:
          dr = DebitTransaction(t_type="DR", total= debit_amount, date=now, description="استرجاع قيمة فاتورة رقم  "+ str(p.id) , invoice_id=invoice_id, current_balance = account.balance)
          db.session.add(dr)
        p.is_refunded = True
        db.session.commit()
        flash(u'تم اعادة مبلغ الفاتورة', 'success')
        return redirect(url_for('viewInvoices'))
  db.session.commit()        
  list1 = p.products
  print("Purchased products: ")
  print(p.products)
  print(type(p.products))
  y = json.loads(p.products) 
  print(y)     
  list_of_items = []
  for key, value in y.items():
    #temp = [key]
    temp = [key,value]
    list_of_items.append(temp) 
  print("Purchased products as a list: ")
  print(list_of_items)
  print()
  for i in list_of_items: 
    print("NOW NOW NOW")
    print(i[0])
    print(i[1][0])
    #form.refund_products.choices += [i[0] + "-" + "SAR " + str(i[1][0])]
    form.refund_products.choices += [i[0]]

  return render_template('refund-invoice.html',form=form, invoice_id=invoice_id, user=u)




def paymentLog(total):
  print("Inside payment log func")
  total = session.get('sumOfCart')
  year = date.today().year
  month = date.today().month
  print(total)
  print(year)
  print(month)
  print(type(year))
  print(type(total))

  payment = Paymentlog(amount=total, year=year,month = month, source="Online")
  db.session.add(payment)
  #plan = Plan4(customer_mobile=mobile, customer_name = name,  start_date=date, p_type = plan_type, is_active = "Active", expiry_date= date2, number_of_pauses = 0, customer = customer)
  #db.session.add(plan)
  db.session.commit()


def revenue(month):
  today = datetime.now()
  #print("Current Year")
  #print(today.year)
  #print("Month Range")
  range_of_days = monthrange(today.year, month)
  #print(range_of_days)
  #print(range_of_days[0])
  #print(range_of_days[1])
  start = date(year=today.year,month=month,day=1)
  end = date(year=today.year,month=month,day=range_of_days[1])

  crs = RevenueTransaction.query.filter(CreditTransaction.date <= end).filter(RevenueTransaction.date >= start).all()
  #session.query(func.avg(Rating.field2).label('average')).filter(Rating.url==url_string.netloc)
  #crs = CreditTransaction.query(func.sum(CreditTransaction.field2).label('average')).filter(CreditTransaction.date <= end).filter(CreditTransaction.date >= start).all()
  
  ######drs = DebitTransaction.query.filter(DebitTransaction.date <= end).filter(DebitTransaction.date >= start).all() 

  #crs.extend(drs)
  #print("CR & DR together")
  #print(crs) 

  summ=0
  for cr in crs: 
      summ += cr.total
  print("Total Revenue is: ")
  print(summ)

  return summ  


@app.route("/show-invoice/<invoice_id>")
@login_required
def showInvoice(invoice_id):
  p = Inv.query.filter_by(id=invoice_id).first()
  if p.is_expense:
    dr = DebitTransaction.query.filter_by(invoice_id=invoice_id).first()
    return render_template('invoice2.html', total = p.total, description = dr.description)
  print("Invoice Data: ")
  print(p.products);
  print(type(p.products))
  print(p.inv_type)
  print(p.status)
  print(p.remaining_balance)
  curr_customer = None
  if p.customer_id != None:
    print(p.customer_id)
    curr_customer = Customer.query.filter_by(id=p.customer_id).first().name
    print(curr_customer)
  else:
    print("customer_id is None")  
  list1 = json.loads(p.products) 
  list_of_items = []
  total = 0
  for key, value in list1.items():
    temp = [key,value]
    list_of_items.append(temp)
  print("::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::")  
  print("++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
  print("::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::")
  print(p.inv_type)

  #print("Current balance before update is ... +++++++++++++++++++++++++++++")
  #print(account.balance)
  now = datetime.now()
  print("Date & Time now")
  print(now)
  now = date.today()
  print(now)
  print(p.inv_type)
  return render_template('invoice.html',products=list_of_items, length= len(list_of_items), total = p.total, vat=p.vat_value, vat_percentage = p.vat_percentage, customer=curr_customer, remaining_balance= None, category = p.inv_type, date=now, invoice_id=invoice_id)

@app.route("/invoice/<invoice_id>")
@login_required
def invoice(invoice_id):
  #list1 = session.get('cart')
  #list_of_items = []
  #print(list1)
  #for key, value in list1.items():
    #temp = [key,value]
    #list_of_items.append(temp)
  #print("CART CONTENT: ")
  #print(list_of_items)
  revenue_total = 0
  revenue = 0  
  paid = 0
  p = Inv.query.filter_by(id=invoice_id).first()
  if p.is_expense:
    dr = DebitTransaction.query.filter_by(invoice_id=invoice_id).first()
    return render_template('invoice2.html', total = p.total, description = dr.description)
  print("Invoice Data: ")
  print(p.products);
  print(type(p.products))
  print(p.inv_type)
  print(p.status)
  print(p.remaining_balance)
  curr_customer = None
  if p.customer_id != None:
    print(p.customer_id)
    curr_customer = Customer.query.filter_by(id=p.customer_id).first().name
    print(curr_customer)
  else:
    print("customer_id is None")  
  list1 = json.loads(p.products) 
  list_of_items = []
  total = 0
  for key, value in list1.items():
    temp = [key,value]
    list_of_items.append(temp)
  print("::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::")  
  print("++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
  print("::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::")
  print(p.category)
  for item in list_of_items: 
    print(item)
    print(item[0])
    print(item[1])
    print(item[1][0])
    print(item[1][1])

    product = Product.query.filter_by(name=item[0]).first()
    paid = item[1][0] * item[1][1]
    if p.category == "تجزئة" or p.category == "Single":
      print("Single")
      revenue = product.single_expense
    elif p.category == "جملة الجملة" or p.category == "Bulk Bulk":
      print("Bulk Bulk")
      revenue = product.bulk_bulk_expense
    elif p.category == "جملة" or p.category == "Bulk":
      print("Bulk")
      revenue = product.bulk_expense
    else:
      print("INOVICE TYPE IS NON OF THE ABOVE")  

    print("Expense is: ")
    print(revenue)
    print(product.single_expense)
    print(paid)
    revenue = revenue * item[1][1]
    print(revenue)
    revenue = paid - revenue
    print(revenue)
    revenue_total += revenue
    print(paid)
    print(revenue)
    print(revenue_total)
  account = Account.query.filter_by(id=1).first()
  vat_account = Account.query.filter_by(id=2).first()
  revenue_account = Account.query.filter_by(id=3).first()
  if p.inv_type != "Loan":
    now = datetime.now()
    revenue_account.balance = revenue_account.balance + revenue_total 
    rev_transaction = RevenueTransaction(total=revenue_total, date=now, description="ارباح فاورة رقم " + str(p.id), invoice_id= p.id, current_balance = revenue_account.balance)
    db.session.add(rev_transaction)
    db.session.commit()
  print("Current balance before update is ... +++++++++++++++++++++++++++++")
  print(account.balance)
  now = datetime.now()
  print("Date & Time now")
  print(now)
  #cr = CreditTransaction(t_type="CR", total=44, date=now, description="selling")
  if p.inv_type == "Cash" or p.inv_type == "كبس":
    print("INV is CASH .. CR is happening")
    print("total minus vat value in cash is")
    total_minus_vat = p.total - p.vat_value
    print(total_minus_vat)
    account.balance = account.balance + total_minus_vat
    vat_account.balance = vat_account.balance + p.vat_value
    db.session.commit()
  #db.session.add(cr)
    cr = CreditTransaction(t_type="CR", total=total_minus_vat, date=now, description="selling", invoice_id=invoice_id, current_balance = account.balance)
    #cr = CreditTransaction(t_type="CR", total=p.total, date=now, description="selling", invoice_id=invoice_id, current_balance = account.balance)
    db.session.add(cr)
    db.session.commit()
    db.session.flush()
    print("Current balance after update is ... +++++++++++++++++++++++++++++")
    print(account.balance)
    print("LAST COMMITED CR Transaction")
    print(cr.id)
    return redirect(url_for('showInvoice', invoice_id=invoice_id))  
  elif p.inv_type == "Loan": 
    print("INV is LOAN .. No transaction is happening")
    print("Current balance AFTER update is ... +++++++++++++++++++++++++++++")
    print(account.balance)  
  #return invoice_id;
    return redirect(url_for('showInvoice', invoice_id=invoice_id))
  elif p.inv_type == "شبكة":
    print("INV is CASH .. CR is happening")
    print("total minus vat value in cash is")
    total_minus_vat = p.total - p.vat_value
    print(total_minus_vat)
    account.balance = account.balance + total_minus_vat
    vat_account.balance = vat_account.balance + p.vat_value
    db.session.commit()
  #db.session.add(cr)
    account.balance = account.balance - 0.80
    #total_credit = p.total - 0.80
    total_credit = total_minus_vat - 0.80
    cr = CreditTransaction(t_type="CR", total=total_credit, date=now, description="selling", invoice_id=invoice_id, current_balance = account.balance)
    db.session.add(cr)
    db.session.commit()
    db.session.flush()
    print("Current balance after update is ... +++++++++++++++++++++++++++++")
    print(account.balance)
    print("LAST COMMITED CR Transaction")
    print(cr.id)
    #return redirect('/show-invoice')
  return redirect(url_for('showInvoice', invoice_id=invoice_id))
  #return render_template('invoice.html',products=list_of_items, length= len(list_of_items), total = p.total, vat=p.vat_value, vat_percentage = p.vat_percentage, customer=curr_customer, remaining_balance= None, category = p.category)

@app.route("/view-loans.html")
@login_required
def viewLoans():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  invoices = Inv.query.filter_by(inv_type="Loan", status="Not Paid").all()
  print("All Loan invoices")
  print(invoices)
  productstore = Inv.query.filter_by(inv_type="Loan", status="Not Paid").join(Customer, Inv.customer_id==Customer.id).all()
  print("The join ...")
  print(productstore)
  #print(productstore[0].customer.name)
  return render_template('view-loans.html', invoices=productstore, len = len(productstore), user=u)


@app.route("/loan-sadad/<customer_id>" , methods=['GET', 'POST'])
@login_required
def sadadLoans1(customer_id):
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  form = Sadad2(request.form)
  sadad_amount = 0
  if form.submit.data:
    pay_amount = form.pay_amount.data
    print("Entered Amount")
    print(pay_amount)
    current_account = Account.query.filter_by(id=1).first()
    vat_account = Account.query.filter_by(id=2).first()
    revenue_account = Account.query.filter_by(id=3).first()

    vat_value = VAT.query.all()
    vat_value = vat_value[0].vat

    new_vat  = vat_value  * float(pay_amount) / 100
    print("VAT amount: ")
    print(new_vat)

    paid_amount = float(pay_amount) - new_vat
    print("Paid amount minus vat")
    print(paid_amount)

    current_account.balance = current_account.balance + paid_amount
    revenue_account.balance = revenue_account.balance + paid_amount
    vat_account.balance = vat_account.balance + new_vat

    cust = Customer.query.filter_by(id=customer_id).first()
    cust.remaining_balance = cust.remaining_balance - float(pay_amount)

    db.session.commit()

    now = datetime.now()

    loan_transaction = LoanTransaction(total = pay_amount, date = now, description="تسديد دين", current_balance = cust.remaining_balance, customer_id = customer_id)
    rev_transaction = RevenueTransaction(total=paid_amount, date=now, description="ارباح فاورة رقم ", invoice_id= None, current_balance = revenue_account.balance)

    db.session.add(loan_transaction)
    db.session.add(rev_transaction)
    db.session.commit()

    print("Loan Transactions are: ")
    print(LoanTransaction.query.all())


    print("ok")
    flash(u'تمت التسديد', 'success')
    return render_template('loan-sadad.html', form = form, customer_id = customer_id, user=u)
  cust = Customer.query.filter_by(id=customer_id).first()  
  form.customer_id.data = customer_id
  form.remianing_balance.data = cust.remaining_balance
  return render_template('loan-sadad.html', form = form, customer_id = customer_id, user=u)



@app.route("/loan-sadad-existing-records" , methods=['GET', 'POST'])
@login_required
def sadadLoans2():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  form = Sadad33(request.form)
  sadad_amount = 0
  if form.submit.data:
    cust = Customer.query.filter_by(name=form.autocomp.data).first()
    customer_id = cust.id
    pay_amount = form.pay_amount.data
    print("Entered Amount")
    print(pay_amount)
    current_account = Account.query.filter_by(id=1).first()
    vat_account = Account.query.filter_by(id=2).first()
    revenue_account = Account.query.filter_by(id=3).first()

    vat_value = VAT.query.all()
    vat_value = vat_value[0].vat

    new_vat  = vat_value  * float(pay_amount) / 100
    print("VAT amount: ")
    print(new_vat)

    paid_amount = float(pay_amount) - new_vat
    print("Paid amount minus vat")
    print(paid_amount)

    current_account.balance = current_account.balance + paid_amount
    revenue_account.balance = revenue_account.balance + paid_amount
    vat_account.balance = vat_account.balance + new_vat


    db.session.commit()

    now = datetime.now()

    loan_transaction = LoanTransaction(total = pay_amount, date = now, description="سداد ذمم سنة 2020", current_balance = cust.remaining_balance, customer_id = customer_id)
    rev_transaction = RevenueTransaction(total=paid_amount, date=now, description="ارباح فاورة رقم ", invoice_id= None, current_balance = revenue_account.balance)

    db.session.add(loan_transaction)
    db.session.add(rev_transaction)
    db.session.commit()

    print("Loan Transactions are: ")
    print(LoanTransaction.query.all())


    print("ok")
    flash(u'تمت التسديد', 'success')
    return render_template('loan-sadad-existing-records.html', form = form, user=u)
  return render_template('loan-sadad-existing-records.html', form = form, user=u)



@app.route("/sadad-loans/<invoice_id>" , methods=['GET', 'POST'])
@login_required
def sadadLoans(invoice_id):
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  form = Sadad(request.form)
  sadad_amount = 0
  if form.submit.data:
    print("SADAD Button is clicked")
    invoices = Inv.query.filter_by(id=invoice_id).first()
    account = Account.query.filter_by(id=1).first()
    #print("Account Balance before Sadad: ")
    #print(account.balance)
    #account.balance = account.balance + float(form.pay_amount.data)
    #db.session.commit()
    #balance = account.balance + float(form.pay_amount.data)

    #if invoices.remaining_balance == 0:
    vat_value = VAT.query.all()
    vat_value = vat_value[0].vat
    print("&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&")
    #print("VAT NOT SUB")
    #print(vat_value[0])
    sadad_percentage = float(form.pay_amount.data) / float(invoices.total)
    print("SADAD Percentage")
    print(sadad_percentage)
    vat = invoices.vat_value * (sadad_percentage)
    print("VAT")
    print(vat)
    sadad_amount = float(form.pay_amount.data) - vat
    print("SADAD Amount")
    print(sadad_amount)
    vat_account = Account.query.filter_by(id=2).first()
    vat_account.balance = vat_account.balance + vat
    account.balance = account.balance + sadad_amount
    invoices.remaining_balance = invoices.remaining_balance -  (sadad_amount + vat)
    db.session.commit()
    y = json.loads(invoices.products)
    if invoices.remaining_balance == 0:
      list_of_items = []
      for key, value in y.items():
        #temp = [key]
        temp = [key,value]
        list_of_items.append(temp)
      paid = 0
      revenue = 0
      revenue_total = 0       
      for item in list_of_items: 
        print(item)
        print(item[0])
        print(item[1])
        print(item[1][0])
        print(item[1][1])

        product = Product.query.filter_by(name=item[0]).first()
        paid = item[1][0] * item[1][1]
        if invoices.category == "تجزئة":
          print("Single")
          revenue = product.single_expense
        elif invoices.category == "جملة الجملة":
          print("Bulk Bulk")
          revenue = product.bulk_bulk_expense
        elif invoices.category == "جملة":
          print("Bulk")
          revenue = product.bulk_expense
        else:
          print("INOVICE TYPE IS NON OF THE ABOVE")  

        print(paid)
        revenue = revenue * item[1][1]
        print(revenue)
        revenue = paid - revenue
        print(revenue)
        revenue_total += revenue
        print(paid)
        print(revenue)
        print(revenue_total)
      revenue_account = Account.query.filter_by(id=3).first()
      revenue_account.balance = revenue_account.balance + revenue_total  
      print("Purchased products as a list: ")
      print(list_of_items)

      db.session.commit() 
    now = datetime.now()
    cr = CreditTransaction(t_type="CR", total=float(sadad_amount), date=now, description="تسديد ذمة", invoice_id=invoice_id, current_balance = account.balance)
    dr = DebitTransaction(t_type="DR", total=float(vat), date=now, description="نقل مبلغ الضريبة لحساب الضريبة", invoice_id=invoice_id, current_balance = account.balance)
    db.session.add(cr)
    db.session.add(dr)
    db.session.commit()
    flash(u'|تم التسديد', 'success')
    return render_template('sadad-loans.html', form = form, invoice_id = invoice_id, user=u)

  invoices = Inv.query.filter_by(id=invoice_id).first()
  form.invoice_id.data = invoice_id
  form.remianing_balance.data = invoices.remaining_balance
  return render_template('sadad-loans.html', form = form, invoice_id = invoice_id, user=u)

@app.route("/sell-branch-1.html")
@login_required
def sellBranchOne():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  form = SearchForm(request.form)
  products = [];
  print('products: ')
  print('HHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHH');
  print(len(products))
  print(products)
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  return render_template('sell-branch-1.html', form=form, products=products, user=u)

@app.route("/sell-branch-2.html")
@login_required
def sellBranchTwo():
  form = SearchForm(request.form)
  products = [];
  print('products: ')
  print('HHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHH');
  print(len(products))
  print(products)
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  return render_template('sell-branch-2.html', form=form, products=products, user=u)


@app.route("/sell-branch-1-cash.html")
@login_required
def sellBranchOneCash():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  if 'cart' not in session:
    print("$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$")
    print("Creating session")
    session['cart'] = {}
  session['cart'] = {}  
  products = [];
  length = len(products);
  form = SellCash(request.form)
  list1 = session.get('cart')
  list_of_items = []
  print(list1)
  if list1 == None:
    list_of_items = []
    return render_template('sell-branch-1-cash.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  for key, value in list1.items():
    temp = [key,value]
    list_of_items.append(temp)
  return render_template('sell-branch-1-cash.html', form=form, products = list1.items(), length = len(list_of_items),user=u);
  #return render_template('sell-branch-1-cash.html', form=form, products=products, length=length)

@app.route("/sell-branch-1-loan.html")
@login_required
def sellBranchOneLoan():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  if 'cart' not in session:
    print("$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$")
    print("Creating session")
    session['cart'] = {}
  session['cart'] = {}  
  products = [];
  length = len(products);
  form = SellLoan(request.form)
  list1 = session.get('cart')
  list_of_items = []
  print(list1)
  if list1 == None:
    list_of_items = []
    return render_template('sell-branch-1-loan.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  for key, value in list1.items():
    temp = [key,value]
    list_of_items.append(temp)
  return render_template('sell-branch-1-loan.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  #return render_template('sell-branch-1-cash.html', form=form, products=products, length=length)  


@app.route("/sell-branch-1-card.html")
@login_required
def sellBranchOneCard():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  if 'cart' not in session:
    print("$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$")
    print("Creating session")
    session['cart'] = {}
  session['cart'] = {}  
  products = [];
  length = len(products);
  form = SellCash(request.form)
  list1 = session.get('cart')
  list_of_items = []
  print(list1)
  if list1 == None:
    list_of_items = []
    return render_template('sell-branch-1-card.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  for key, value in list1.items():
    temp = [key,value]
    list_of_items.append(temp)
  return render_template('sell-branch-1-card.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  #return render_template('sell-branch-1-cash.html', form=form, products=products, length=length)

@app.route("/sell-branch-1-kabs.html")
@login_required
def sellBranchOneKabs():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  if 'cart' not in session:
    print("$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$")
    print("Creating session")
    session['cart'] = {}
  session['cart'] = {}  
  products = [];
  length = len(products);
  form = SellCash(request.form)
  list1 = session.get('cart')
  list_of_items = []
  print(list1)
  if list1 == None:
    list_of_items = []
    return render_template('sell-branch-1-kabs.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  for key, value in list1.items():
    temp = [key,value]
    list_of_items.append(temp)
  return render_template('sell-branch-1-kabs.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  #return render_template('sell-branch-1-cash.html', form=form, products=products, length=length)




@app.route("/sell-branch-2-cash.html")
@login_required
def sellBranchTwoCash():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  if 'cart' not in session:
    print("$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$")
    print("Creating session")
    session['cart'] = {}
  session['cart'] = {}  
  products = [];
  length = len(products);
  form = SellCash(request.form)
  list1 = session.get('cart')
  list_of_items = []
  print(list1)
  if list1 == None:
    list_of_items = []
    return render_template('sell-branch-2-cash.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  for key, value in list1.items():
    temp = [key,value]
    list_of_items.append(temp)
  return render_template('sell-branch-2-cash.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  #return render_template('sell-branch-1-cash.html', form=form, products=products, length=length)

@app.route("/sell-branch-2-loan.html")
@login_required
def sellBranchTwoLoan():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  if 'cart' not in session:
    print("$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$")
    print("Creating session")
    session['cart'] = {}
  session['cart'] = {}  
  products = [];
  length = len(products);
  form = SellLoan(request.form)
  list1 = session.get('cart')
  list_of_items = []
  print(list1)
  if list1 == None:
    list_of_items = []
    return render_template('sell-branch-2-loan.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  for key, value in list1.items():
    temp = [key,value]
    list_of_items.append(temp)
  return render_template('sell-branch-2-loan.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  #return render_template('sell-branch-1-cash.html', form=form, products=products, length=length)  


@app.route("/sell-branch-2-card.html")
@login_required
def sellBranchTwoCard():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  if 'cart' not in session:
    print("$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$")
    print("Creating session")
    session['cart'] = {}
  session['cart'] = {}  
  products = [];
  length = len(products);
  form = SellCash(request.form)
  list1 = session.get('cart')
  list_of_items = []
  print(list1)
  if list1 == None:
    list_of_items = []
    return render_template('sell-branch-2-card.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  for key, value in list1.items():
    temp = [key,value]
    list_of_items.append(temp)
  return render_template('sell-branch-2-card.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  #return render_template('sell-branch-1-cash.html', form=form, products=products, length=length)

@app.route("/sell-branch-2-kabs.html")
@login_required
def sellBranchTwoKabs():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  if 'cart' not in session:
    print("$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$")
    print("Creating session")
    session['cart'] = {}
  session['cart'] = {}  
  products = [];
  length = len(products);
  form = SellCash(request.form)
  list1 = session.get('cart')
  list_of_items = []
  print(list1)
  if list1 == None:
    list_of_items = []
    return render_template('sell-branch-2-kabs.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  for key, value in list1.items():
    temp = [key,value]
    list_of_items.append(temp)
  return render_template('sell-branch-2-kabs.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  #return render_template('sell-branch-1-cash.html', form=form, products=products, length=length)



class Items:
  product_name = ""
  product_id = 0

@app.route('/selectform', methods=['POST'])
def updateselect():
    print("INSIDE SELECT FORM FUNC")
    deviceAmount = request.form.get('inv_category')
    print(deviceAmount)
    #choices = [('device{}'.format(i), i) for i in range(deviceAmount)]
    myAnswer = "Ok"
    #response = make_response(json.dumps(myAnswer))
    response = json.dumps(myAnswer)
    #response.content_type = 'application/jsons'
    return response
def reCalculateInvoice(id):
    now = datetime.now()
    print("Inside reCalculateInvoice")
    print("### Changes begin ###")

    inv = Inv.query.filter_by(id=id).first()
    list1=json.loads(inv.products)
    list_of_items = []
    for key, value in list1.items():
      temp = [key,value]
      list_of_items.append(temp)
      branchProduct = BranchOneProduct.query.filter_by(name=key).first()
      branchProduct.quantity = branchProduct.quantity - int(value[1])
      db.session.commit()
    y = json.dumps(list1)
    print(y)
    total = 0
    for item in list_of_items:
      total = total + (item[1][0] * item[1][1])
    print("The total of invoice nad transactions is ...")
    print(total)

    vat_value = VAT.query.all()
    vat_value = vat_value[0].vat
    vat = vat_value * total / 100


    current_account = Account.query.filter_by(id=1).first()
    vat_account = Account.query.filter_by(id=2).first()

    if vat > inv.vat_value:
      diff = vat-inv.vat_value
      current_account.balance = current_account.balance - diff
      vat_account.balance = vat_account.balance + diff
      db.session.commit()
      #dr = DebitTransaction(t_type="DR", total=diff, date=now, description="تعديل فاتورة رقم " + str(id), invoice_id=id, current_balance = current_account.balance)
      #cr = CreditTransaction(t_type="CR", total=diff, date=now, description="تعديل فاتورة رقم " + str(id) + "ايداع لحساب الضريبية", invoice_id=id, current_balance = current_account.balance)
      #db.session.add(dr)
      #db.session.add(cr)
      db.session.commit()
    elif vat < inv.vat_value:
      diff = inv.vat_value - vat
      #current_account.balance = current_account.balance + diff
      vat_account.balance = vat_account.balance - diff
      db.session.commit()
      #dr = DebitTransaction(t_type="DR", total=diff, date=now, description="تعديل فاتورة رقم " + str(id) + "خصم حساب الضريبة", invoice_id=id, current_balance = current_account.balance)
      #cr = CreditTransaction(t_type="CR", total=diff, date=now, description="تعديل فاتورة رقم " + str(id), invoice_id=id, current_balance = current_account.balance)
      #db.session.add(dr)
      #db.session.add(cr)
      db.session.commit()

    inv.total = total + vat
    inv.vat_value = vat
    inv.is_modified = True
    #setRevenue(inv.id, )
    db.session.commit()
    print("Total Invoice: ")
    print(inv.total)
    print("Total VTA: ")
    print(inv.vat_value)
    print("### Changes end ###")
    return 0


def setRevenue(id, name, price, quantity, old_price, old_quantity):
  print("Inside setRevenue() func.... ")
  p= Inv.query.filter_by(id=id).first()
  product = Product.query.filter_by(name=name).first()
  expense = 0
  revenue = 0
  if p.category == "تجزئة":
    print("Single")
    expense = product.single_expense
  elif p.category == "جملة الجملة":
    print("Bulk Bulk")
    expense = product.bulk_bulk_expense
  elif p.category == "جملة":
    print("Bulk")
    expense = product.bulk_expense
  else:
    print("INOVICE TYPE IS NON OF THE ABOVE")  
  quantity_difference = abs(old_quantity - quantity)
  price_minus_expense = abs(price-expense)
  revenue = quantity_difference * price_minus_expense
  revenue_account = Account.query.filter_by(id=3).first()
  old_total_price = old_price * old_quantity
  new_total_price = price * quantity
  if new_total_price > old_total_price: 
    revenue_account.balance = revenue_account.balance + revenue
  elif new_total_price < old_total_price:
    revenue_account.balance = revenue_account.balance - revenue
  #rev_transaction = RevenueTransaction(total=revenue_total, date=now, description="ارباح فاورة رقم " + str(p.id), invoice_id= p.id, current_balance = revenue_account.balance)  
  db.session.commit()
  print("Reevenue Account Balance: ")
  print(revenue_account.balance)  

def amendChangestoInv(id, name, price, quantity):
    now = datetime.now()
    print("Amending Changes to INV")
    print("## Start of Changes ##")
    inv= Inv.query.filter_by(id=id).first()
    product = None
    print(inv)
    print(inv.products)
    print(type(inv.products))
    print(json.loads(inv.products))
    y=json.loads(inv.products)
    print(y[name])

    ################################################
    final_price = 0
    final_quantity = 0
    print("Inside amendChangestoInv() ")
    print("::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::")
    print("Registered price and quantity")
    print(y[name][0])
    print(y[name][1])
    print("The new price and quantity values")
    print(price)
    print(quantity)
    print("Invoice category: ")
    print(inv.category)
    print("Invoice type: ")
    print(inv.inv_type)
    print("::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::")
    if inv.inv_type != "Loan":
      print("Sure")
      #####setRevenue(id,name,price,quantity,y[name][0],y[name][1])
    # y[name][0] is the registered price
    # y[name][1] is the registered quantity
    if quantity > y[name][1]:
      print("Q > REGISTERED Q")
      total = quantity - y[name][1]
      y[name][1] = y[name][1] - total
      final_quantity = total
    elif quantity < y[name][1]:
      print("Q < REGISTERED Q")
      total =  y[name][1] - quantity
      y[name][1] = y[name][1] + total
      final_quantity = total
    else:
      total_quantity = quantity  


    if price > y[name][0]: 
      print("P > REGISTERED P")
      total_price = price - y[name][0]
      price_for_registered_products = total_price * y[name][1]
      price_for_new = total_price * quantity
      final_price = price_for_registered_products + price_for_new
      final_price = price_for_new
      final_quantity = quantity
      print("CR Transaction amount")
      print(total_price)
      account = Account.query.filter_by(id=1).first()
      vat_account = Account.query.filter_by(id=2).first()

      account.balance = account.balance + final_price
      db.session.commit()
      cr = CreditTransaction(t_type="CR", total=final_price, date=now, description="تعديل فاتورة رقم " + str(id), invoice_id=id, current_balance = account.balance)
      db.session.commit()
    elif price < y[name][0]: 
      print("P < REGISTERED P")
      total_price = y[name][0] - price
      print("Total price")
      print(total_price)
      #total_quantity =  y[name][1] - quantity
      print("Total Quantity")
      print(total_quantity)
      price_to_be_refunded = total_price * total_quantity
      print("DR Transaction amount")
      print(price_to_be_refunded)
      final_price = price_to_be_refunded
      final_quantity = quantity
      account = Account.query.filter_by(id=1).first()
      vat_account = Account.query.filter_by(id=2).first()

      account.balance = account.balance - final_price
      db.session.commit()
      cr = DebitTransaction(t_type="DR", total=price_to_be_refunded, date=now, description="تعديل فاتورة رقم " + str(id), invoice_id=id, current_balance = account.balance)
      db.session.commit()
    else:
        print("P == REGISTERED P")
        final_price = y[name][0] * final_quantity
        print(y[name][0])
        print(y[name][1])
        #final_quantity = y[name][1]




    print("Total price and quantity: ")
    print(final_price)
    print(final_quantity)
    print("Old price and quantity")
    print(y[name])
    print(y[name][0])
    print(y[name][1])
    print("## End of Changes ##")
    ################################################
    current_account = Account.query.filter_by(id=1).first()
    revenue_account = Account.query.filter_by(id=3).first()
    print("Current Account Balance:")
    print(current_account.balance)
    print("Reveneue Account Balance: ")
    print(revenue_account.balance)

    print("Old Values: ")
    print(y[name][0])
    print(y[name][1])

    print("New Values: ")
    print(price)
    print(quantity)

    reCalculateRevenue(id, name, price, quantity, y[name][0],y[name][1])
    y[name][0] = price
    y[name][1] = quantity
    print(y)
    print(json.dumps(y))
    final_version_of_products = json.dumps(y)
    inv.products = final_version_of_products
    db.session.commit()
    reCalculateInvoice(id)
    return redirect(url_for('invoiceEdit', invoice_id=id))



    #if inv.initiator == "Branch1":
      #product = 
    #elif inv.initiator == "Branch2"

def reCalculateRevenue(id, name, price,quantity,old_price,old_quantity):
  print("Inside recalulate Revenue")
  print(id)
  print(name)
  print(price)
  print(quantity)
  product = Product.query.filter_by(name = name).first()
  product_expense = product.single_expense
  price_after_change = price - product_expense
  new_total = price_after_change * quantity
  old_total = old_price * old_quantity
  revenue_account = Account.query.filter_by(id=3).first()
  revenue_account.balance = revenue_account.balance - (old_total - product_expense)
  db.session.commit()
  print("Old total")
  print(old_total)
  print("New total")
  print(new_total)
  revenue_account.balance = revenue_account.balance + new_total
  db.session.commit()

def placeChangesToInvoice(id, name, price, quantity):
    inv= Inv.query.filter_by(id=id).first()
    product = None
    print(inv)
    print(inv.products)
    print(type(inv.products))
    print(json.loads(inv.products))
    y=json.loads(inv.products)
    print(y[name])

    ################################################
    final_price = 0
    final_quantity = 0
    print("Inside placeChangesToInvoice() ")
    print("::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::")
    print("Registered price and quantity")
    print(y[name][0])
    print(y[name][1])
    print("The new price and quantity values")
    print(price)
    print(quantity)
    print("Invoice category: ")
    print(inv.category)
    print("Invoice type: ")
    print(inv.inv_type)
    print("::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::")

    if quantity > y[name][1]: 
      print("New quantity is bigger")
    elif quantity < y[name][1]:   
      print("New quantity is smaller")
    else:
      print("New and old quantities are equal")

    if price > y[name][0]:
      print("New price is greater")
    elif price < y[name][0]:
      print("New price is smaller")
    else:
      print("New and old prices and equal")

    p= Inv.query.filter_by(id=id).first()
    product = Product.query.filter_by(name=name).first()
    expense = 0
    revenue = 0
    if p.category == "تجزئة" or p.category == "Single":
      print("Single")
      expense = product.single_expense
    elif p.category == "جملة الجملة" or p.category == "Bulk Bulk":
      print("Bulk Bulk")
      expense = product.bulk_bulk_expense
    elif p.category == "جملة" or p.category == "Bulk":
      print("Bulk")
      expense = product.bulk_expense
    else:
      print("INOVICE TYPE IS NON OF THE ABOVE")


    ### #### ### ### ### ### ### ###
    if inv.inv_type != "Loan":
      total_old = y[name][0] * y[name][1]
      total_new = price * quantity
      print("Old Total")
      print(total_old)
      print("New Total")
      print(total_new)

      vat_value = VAT.query.all()
      vat_value = vat_value[0].vat

      old_vat = vat_value * total_old / 100
      new_vat = vat_value * total_new / 100
      print("Old VAT")
      print(old_vat)
      print("New VAT")
      print(new_vat)


      #print(y[name][0])
      #print(y[name][1])
      #print(expense)

      old_revenue = abs(total_old - (y[name][1] * expense) )
      new_revenue = abs(total_new - (quantity * expense) )
      print("Old Revenue")
      print(old_revenue)
      print("New Revenue")
      print(new_revenue)

      current_account = Account.query.filter_by(id=1).first()
      vat_account = Account.query.filter_by(id=2).first()
      revenue_account = Account.query.filter_by(id=3).first()

      current_account.balance = current_account.balance - total_old
      vat_account.balance = vat_account.balance - old_vat
      revenue_account.balance =  revenue_account.balance - old_revenue

      db.session.commit()

      current_account.balance = current_account.balance + total_new
      vat_account.balance = vat_account.balance + new_vat
      revenue_account.balance =  revenue_account.balance + new_revenue

      rev_transaction = RevenueTransaction.query.filter_by(invoice_id=id).first()
      rev_transaction.total = new_revenue
      db.session.commit()


    y[name][0] = price
    y[name][1] = quantity
    print(y)
    print(json.dumps(y))
    final_version_of_products = json.dumps(y)
    inv.products = final_version_of_products
    db.session.commit()

    inv = Inv.query.filter_by(id=id).first()
    list1=json.loads(inv.products)
    list_of_items = []
    for key, value in list1.items():
      temp = [key,value]
      list_of_items.append(temp)
      db.session.commit()
    y = json.dumps(list1)
    print(y)
    total = 0
    for item in list_of_items:
      total = total + (item[1][0] * item[1][1])
    print("The total of invoice nad transactions is ...")
    print(total)

    vat_value = VAT.query.all()
    vat_value = vat_value[0].vat

    new_vat_invoice  = vat_value * total / 100
    inv.vat_value = new_vat_invoice
    inv.total = total + new_vat_invoice

    db.session.commit()

@app.route('/loanPayments',methods=['GET', 'POST'])
def loanPayments():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  customers = Customer.query.filter(Customer.remaining_balance > 0, Customer.remaining_balance != None).all()
  return render_template('loan-payments.html', customers=customers, len = len(customers), user=u)



@app.route('/editables', methods=['POST'])
def editables():

    print("INSIDE editables FUNC")
    print("jquey value: ")
    title = request.form.to_dict()
    print(title)
    print(type(title))
    print(title.items())
    print(title['name'])
    print(type(title['price']))

    id = int(title['id'])
    price = float(title['price'])
    quantity = int(title['quantity'])
    name = title['name']
    #####amendChangestoInv(id, name, price,quantity)
    placeChangesToInvoice(id, name, price,quantity)

    #list_of_items = []
    #print(title[' '])
    #for key, value in title.items():
      #temp = [value]
      #print(temp)
      #list_of_items.append(temp)
    #print(list_of_items)  

    #print(title[name])
    #print(title["price"])
    #print(title["quantity"])
    #print(title["id"])


    #deviceAmount = request.form.get('inv_category')
    #print(deviceAmount)
    #data = request.get_json()
    #print("Loooooads of data")
    #print(data)
    #choices = [('device{}'.format(i), i) for i in range(deviceAmount)]
    #response = make_response(json.dumps(myAnswer))
    #response = json.dumps(myAnswer)
    #print("MYYYYYYYYY JQUERY RESPONSE: ")
    #print(response)
    #response.content_type = 'application/jsons'
    #return redirect(url_for('invoiceEdit', invoice_id=id))
    return "Ok"



def sellingHistory(customer_name, product_name):
  print("Inside Selling History func()")
  #c = Customer.query.filter_by(name=form.autocompcustomer.data).first()
  c = Customer.query.filter_by(name=customer_name).first()
  print(c)
  inv = Inv.query.filter_by(customer_id=c.id).all()
  print("Length of Invs is ", len(inv))
  prices =[]
  dates = []
  for i in inv:
    y = json.loads(i.products)
    print(y)
    print(type(y))
    #print(i)
    #print(i.products)
    #print(type(i.products))
    #print("Test Arabic")
    #print("سلام")
    if product_name in y:
      print("It does exist")
      print(y[product_name])
      print(y[product_name][0])
      prices.append(y[product_name][0])
      cr = CreditTransaction.query.filter_by(invoice_id=i.id).all()
      for transaction in cr:
        dates.append(transaction.date)

      #print("Sanity Check...")  
      #print(cr)
      #print(prices)
      #print(dates)

    else:
      print("It does not exist")
  return prices    

@app.route("/history_log", methods=['GET', 'POST'])
@login_required
def historyLog():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  form = SellLoan(request.form)
  list_of_items = []
  prices = []
  print("INSIDE HISTORY LOG FUNC()")
  if form.submit.data:
    prices = sellingHistory(form.autocompcustomer.data, form.autocomp.data)
    print("TOTAL HISTORY")
    print(prices)
    return render_template('history_log.html', form=form, list = prices, len = len(prices), user=u);
  return render_template('history_log.html', form=form, list = prices, len = len(prices), user=u);

@app.route("/sell-branch-1-loan-to-db.html", methods=['GET', 'POST'])
@login_required
def sellBranchOneLoanDB1():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  list_of_items = []
  print("my func!!")
  form = SellLoan(request.form)
  customer_name = form.autocompcustomer.data
  number_of_mine = 1
  if form.get_price.data:
    print("Get Price button is clicked")
    #product_name = 
    branchProduct = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
    print(request.form.get('inv_category'))
    invoice_category = ""
    invoice_category_arabic = ""
    requested_price = 0
    if request.form.get('inv_category') == "جملة " or request.form.get('inv_category')  == "Bulk":
      print("Inside bulk")
      invoice_category = "bulk_price"
      invoice_category_arabic = "جملة "
      requested_price = branchProduct.bulk_price
    elif request.form.get('inv_category') == "جملة  الجملة " or request.form.get('inv_category')  == "Bulk Bulk":
      print("Inside bulk bulk")
      invoice_category = "bulk_bulk_price"
      invoice_category_arabic = "جملة  الجملة "
      requested_price = branchProduct.bulk_bulk_price
    elif request.form.get('inv_category') == "تجزئة " or request.form.get('inv_category')  == "Single":
      print("Inside single")
      invoice_category = "single_price"
      invoice_category_arabic = "تجزئة "
      requested_price = branchProduct.single_price
    else:
      print("Inside Non of them") 
    print("INVOICE CATEGORY: ")
    print(invoice_category)
    print("REQUESTED PRICE")
    print(requested_price)
    #branchProduct = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
    form.price.data = requested_price
    #form.available_quantity.data =  100
    list1 = session.get('cart')
    list_of_items = []
    print(list1)
    if list1 == None:
      list_of_items = []
      return render_template('sell-branch-1-loan.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
    for key, value in list1.items():
      temp = [key,value]
      list_of_items.append(temp)
    form = SellLoan()
    form.price.data = requested_price
    print("FORM DATA PASSES")
    print(form.price.data)
    #form.available_quantity = int(branchProduct.quantity)
    form.autocomp.data = branchProduct.name
    form.autocompcustomer.data = customer_name
    form.inv_category.data = invoice_category_arabic
    return render_template('sell-branch-1-loan.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
    #return "Get Price button is clicked"
  elif form.submit.data:
    print("Submit button is clicked")
    print("Submit button is clicked")      
    print(form.autocomp.data)
    print(form.quantity.data)
    print(form.price.data)
    pb1 = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
    p = Product.query.filter_by(name=form.autocomp.data).first()
    print("Quantity check ...")
    print(form.quantity.data)
    print(type(form.quantity.data))
    #print(type(p.b1_quantity))
    #print(p.b1_quantity)
    if form.quantity.data >= pb1.quantity:
      flash(u'Product Quantity is out of stock', 'danger')
      list1 = session.get('cart')
      list_of_items = []
      print(list1)
      if list1 == None:
        list_of_items = []
        return render_template('sell-branch-1-loan.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
      for key, value in list1.items():
        temp = [key,value]
        list_of_items.append(temp)
      return render_template('sell-branch-1-loan.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
      #p = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
    p_all = BranchOneProduct.query.all()
    print("all products")
    print(p_all)
    print("product")
    ##print(p)
    print("product price: ")
    ##print(p.price)
    print("product availability: ")
    ##print(p.quantity)
    ##print(type(p.quantity))
    print("product quantity")
    print(form.quantity.data)
    print(type(form.quantity.data))
    invoice = Invoice()
    db.session.add(invoice)
    db.session.commit()

    print("Invo query: ")
    print(Invoice.query.all())
    print("First Invoice is >>>")
    invoice = Invoice.query.filter_by(invoice_id=1).first()
    print(invoice)
    print("PRINT ALL PRODUCTS ASSOCIATED WITH THIS INVOICE:")
    print(invoice.products)
    print("THE END!")
    #p = Product(name="switches", price=10,shelf="W100", quantity=1000)
    db.session.add(p)
    db.session.commit()
    invoice.products.append(p)
    db.session.commit()
    print("Equiry joint table for INVOICES")
    ##print(Class.query.join(Class.students).all())
    print(invoice.products)
    now = datetime.now()
    #t = Transaction(t_type="CR", total=1000,date=now, description="buy stuff", balance=45000, p_type="Cash")
    #t.invoice = invoice 
    #db.session.add(t)
    #db.session.commit()
    #print("Transactions : ")
    #print(Transaction.query.filter_by(id=1).first())
    #print(Transaction.query.filter_by(id=1).first().invoice)
    """
    cart_item = {'pineapples': '10', 'apples': '20', 'mangoes': '30'}
    print("HERE I'M ...")
    print(cart_item)
    print(type(cart_item))
    items_list = []
    items_list.append(p)
    list_of_items=session.get('cart')
    session['cart'] = cart_item
    list_of_items=session.get('cart')
    print("DICT VALUES ARE: ....")
    print(session.get('cart'))
    print(len(list_of_items))
    """
    if 'cart' not in session:
      print("$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$")
      print("Creating session")
      session['cart'] = {}

    #session['cart'] = {}
    item = [ float(form.price.data), form.quantity.data]
    if item:
      cart_list = session['cart']
      print("CART: ")
      cart_list[form.autocomp.data] = item
      print(cart_list)
      print(type(cart_list))
      session['cart'] = cart_list
      list1 = session.get('cart')
      list_of_items = []
      print(list1)
      for key, value in list1.items():
        temp = [key,value]
        list_of_items.append(temp)
      print("The Full List: ")
      print(list_of_items)
      print(list_of_items[0][0])
      print(list_of_items[0][1])
      print(list_of_items[0][1][0])
      print(list_of_items[0][1][1])
      print(len(list_of_items))
      #str1 = ''.join(list_of_items)
      print("MY STRING IS")
      #product = Product(name=form.name.data, price=form.price.data,shelf=form.shelf.data, w_quantity=form.quantity.data, b1_quantity=0, b2_quantity=0)
    
      ##y = json.dumps(list1)
      ##print(y)
      ##inv = Inv(products=y)
      ##db.session.add(inv)
      ##db.session.commit()
      ##print(Inv.query.all())
      ##invvv= Inv.query.all()
      ##print(invvv[0].products)



      #products = db.Column(db.varchar(250))
      #db.session.add(product)
      #y = json.dumps(list1)
      #print(y)
      #print(request.form['clear'])
      #if request.form['submit'] is not None:
       # session.pop('cart')
        #print("Cart is cleared")




      #cart_list.append(item)
      #session['cart'] = cart_list  # 
      #session.modified = True
      #print("Shopping cart: ")
      #print(session.get('cart'))
      #print(session['cart'])
      #session['cart'] = items_list
      #p = jsonify(items_list)

    return render_template('sell-branch-1-loan.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
    #return "Submit button is clicked"
  elif form.get_price.data:
    print("Get Price button is clicked")
    #product_name = 
    branchProduct = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
    print(request.form.get('inv_category'))
    invoice_category = ""
    invoice_category_arabic = ""
    requested_price = 0
    if request.form.get('inv_category') == "جملة ":
      invoice_category = "bulk_price"
      invoice_category_arabic = "جملة "
      requested_price = branchProduct.bulk_price
    elif request.form.get('inv_category') == "جملة  الجملة ":
      invoice_category = "bulk_bulk_price"
      invoice_category_arabic = "جملة  الجملة "
      requested_price = branchProduct.bulk_bulk_price
    elif request.form.get('inv_category') == "تجزئة ":
      invoice_category = "single_price"
      invoice_category_arabic = "تجزئة "
      requested_price = branchProduct.single_price
    print("INVOICE CATEGORY: ")
    print(invoice_category)
    print("REQUESTED PRICE")
    print(requested_price)
    #branchProduct = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
    form.price.data = requested_price
    #form.available_quantity.data =  100
    list1 = session.get('cart')
    list_of_items = []
    print(list1)
    if list1 == None:
      list_of_items = []
      return render_template('sell-branch-1-loan.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
    for key, value in list1.items():
      temp = [key,value]
      list_of_items.append(temp)
    form = SellLoan()
    form.price.data = requested_price
    print("FORM DATA PASSES")
    print(form.price.data)
    #form.available_quantity = int(branchProduct.quantity)
    form.autocomp.data = branchProduct.name
    form.autocompcustomer.data = customer_name
    form.inv_category.data = invoice_category_arabic
    return render_template('sell-branch-1-loan.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
    #return invoice_category + str(requested_price)
  elif form.confirm.data:
    list1 = session.get('cart')
    list_of_items = []
    for key, value in list1.items():
      temp = [key,value]
      list_of_items.append(temp)
      branchProduct = BranchOneProduct.query.filter_by(name=key).first()
      branchProduct.quantity = branchProduct.quantity - int(value[1])
      db.session.commit()
    y = json.dumps(list1)
    print(y)
    total = 0
    for item in list_of_items:
      total = total + (item[1][0] * item[1][1])
    print("The total of invoice nad transactions is ...")
    print(total)
    #(percent * whole) / 100.0
    #vat_percentage = VAT(vat = 15)
    #db.session.add(vat_percentage)
    #db.session.commit()
    vat_value = VAT.query.all()
    #print(vat_value)
    #print(vat_value[0])
    #print(vat_value[0].vat)
    vat_value = vat_value[0].vat
    print("&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&")
    #print("VAT NOT SUB")
    #print(vat_value[0])
    vat = vat_value * total / 100
    total = total + vat
    print("The list..")
    print(list_of_items)
    print("The total is ...")
    print(total)
    print(form.autocompcustomer.data)
    c = Customer.query.filter_by(name=form.autocompcustomer.data).first()
    percentage = VAT.query.filter_by(id=1).first()
    print(percentage)
    print(percentage.vat)
    now = datetime.now()
    inv = Inv(products=y, vat_value=vat, vat_percentage= percentage.vat,  total=total, inv_type = "Loan", status = "Not Paid", remaining_balance=total, initiator = "Branch1" ,category= request.form.get('inv_category'), date = now, customer_id = c.id)
    db.session.add(inv)
    db.session.commit()
    db.session.flush()
    print("LAST COMMITED INVOICE")
    print(inv.id)
    print(Inv.query.all())
    invvv= Inv.query.all()
    print(invvv[0].products)
    c.remaining_balance = c.remaining_balance + total
    db.session.commit()
    return redirect(url_for('invoice', invoice_id=inv.id))  
  else:
    list1 = session.get('cart')
    list_of_items = []
    print(list1)
    for key, value in list1.items():
      temp = [key,value]
      list_of_items.append(temp)   
    print("CART CONTENT")
    print(list_of_items)  
    return render_template('sell-branch-1-loan.html', form=form, products = list_of_items, length = len(list_of_items), user=u);  
    #return "Finished"

  list1 = session.get('cart')
  list_of_items = []
  print(list1)
  for key, value in list1.items():
    temp = [key,value]
    list_of_items.append(temp)   
  print("CART CONTENT")
  print(list_of_items)  
  return render_template('sell-branch-1-loan.html', form=form, products = list_of_items, length = len(list_of_items), user=u);  
  #return "Finished"       




@app.route("/sell-branch-2-loan-to-db.html", methods=['GET', 'POST'])
@login_required
def sellBranchTwoLoanDB1():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  list_of_items = []
  print("my func!!")
  form = SellLoan(request.form)
  customer_name = form.autocompcustomer.data
  number_of_mine = 1
  if form.get_price.data:
    print("Get Price button is clicked")
    #product_name = 
    branchProduct = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
    print(request.form.get('inv_category'))
    invoice_category = ""
    invoice_category_arabic = ""
    requested_price = 0
    if request.form.get('inv_category') == "جملة " or request.form.get('inv_category')  == "Bulk":
      print("Inside bulk")
      invoice_category = "bulk_price"
      invoice_category_arabic = "جملة "
      requested_price = branchProduct.bulk_price
    elif request.form.get('inv_category') == "جملة  الجملة " or request.form.get('inv_category')  == "Bulk Bulk":
      print("Inside bulk bulk")
      invoice_category = "bulk_bulk_price"
      invoice_category_arabic = "جملة  الجملة "
      requested_price = branchProduct.bulk_bulk_price
    elif request.form.get('inv_category') == "تجزئة " or request.form.get('inv_category')  == "Single":
      print("Inside single")
      invoice_category = "single_price"
      invoice_category_arabic = "تجزئة "
      requested_price = branchProduct.single_price
    else:
      print("Inside Non of them") 
    print(invoice_category)
    print("REQUESTED PRICE")
    print(requested_price)
    #branchProduct = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
    form.price.data = requested_price
    #form.available_quantity.data =  100
    list1 = session.get('cart')
    list_of_items = []
    print(list1)
    if list1 == None:
      list_of_items = []
      return render_template('sell-branch-2-loan.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
    for key, value in list1.items():
      temp = [key,value]
      list_of_items.append(temp)
    form = SellLoan()
    form.price.data = requested_price
    print("FORM DATA PASSES")
    print(form.price.data)
    #form.available_quantity = int(branchProduct.quantity)
    form.autocomp.data = branchProduct.name
    form.autocompcustomer.data = customer_name
    form.inv_category.data = invoice_category_arabic
    return render_template('sell-branch-2-loan.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
    #return "Get Price button is clicked"
  elif form.submit.data:
    print("Submit button is clicked")
    print("Submit button is clicked")      
    print(form.autocomp.data)
    print(form.quantity.data)
    print(form.price.data)
    pb1 = BranchTwoProduct.query.filter_by(name=form.autocomp.data).first()
    p = Product.query.filter_by(name=form.autocomp.data).first()
    print("Quantity check ...")
    print(form.quantity.data)
    print(type(form.quantity.data))
    #print(type(p.b1_quantity))
    #print(p.b1_quantity)
    if form.quantity.data >= pb1.quantity:
      flash(u'Product Quantity is out of stock', 'danger')
      list1 = session.get('cart')
      list_of_items = []
      print(list1)
      if list1 == None:
        list_of_items = []
        return render_template('sell-branch-2-loan.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
      for key, value in list1.items():
        temp = [key,value]
        list_of_items.append(temp)
      return render_template('sell-branch-2-loan.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
      #p = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
    p_all = BranchTwoProduct.query.all()
    print("all products")
    print(p_all)
    print("product")
    ##print(p)
    print("product price: ")
    ##print(p.price)
    print("product availability: ")
    ##print(p.quantity)
    ##print(type(p.quantity))
    print("product quantity")
    print(form.quantity.data)
    print(type(form.quantity.data))
    invoice = Invoice()
    db.session.add(invoice)
    db.session.commit()

    print("Invo query: ")
    print(Invoice.query.all())
    print("First Invoice is >>>")
    invoice = Invoice.query.filter_by(invoice_id=1).first()
    print(invoice)
    print("PRINT ALL PRODUCTS ASSOCIATED WITH THIS INVOICE:")
    print(invoice.products)
    print("THE END!")
    #p = Product(name="switches", price=10,shelf="W100", quantity=1000)
    db.session.add(p)
    db.session.commit()
    invoice.products.append(p)
    db.session.commit()
    print("Equiry joint table for INVOICES")
    ##print(Class.query.join(Class.students).all())
    print(invoice.products)
    now = datetime.now()

    if 'cart' not in session:
      print("$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$")
      print("Creating session")
      session['cart'] = {}

    #session['cart'] = {}
    item = [ float(form.price.data), form.quantity.data]
    if item:
      cart_list = session['cart']
      print("CART: ")
      cart_list[form.autocomp.data] = item
      print(cart_list)
      print(type(cart_list))
      session['cart'] = cart_list
      list1 = session.get('cart')
      list_of_items = []
      print(list1)
      for key, value in list1.items():
        temp = [key,value]
        list_of_items.append(temp)
      print("The Full List: ")
      print(list_of_items)
      print(list_of_items[0][0])
      print(list_of_items[0][1])
      print(list_of_items[0][1][0])
      print(list_of_items[0][1][1])
      print(len(list_of_items))
      #str1 = ''.join(list_of_items)
      print("MY STRING IS")


    return render_template('sell-branch-2-loan.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
    #return "Submit button is clicked"
  elif form.get_price.data:
    print("Get Price button is clicked")
    #product_name = 
    branchProduct = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
    print(request.form.get('inv_category'))
    invoice_category = ""
    invoice_category_arabic = ""
    requested_price = 0
    if request.form.get('inv_category') == "جملة " or request.form.get('inv_category')  == "Bulk":
      print("Inside bulk")
      invoice_category = "bulk_price"
      invoice_category_arabic = "جملة "
      requested_price = branchProduct.bulk_price
    elif request.form.get('inv_category') == "جملة  الجملة " or request.form.get('inv_category')  == "Bulk Bulk":
      print("Inside bulk bulk")
      invoice_category = "bulk_bulk_price"
      invoice_category_arabic = "جملة  الجملة "
      requested_price = branchProduct.bulk_bulk_price
    elif request.form.get('inv_category') == "تجزئة " or request.form.get('inv_category')  == "Single":
      print("Inside single")
      invoice_category = "single_price"
      invoice_category_arabic = "تجزئة "
      requested_price = branchProduct.single_price
    else:
      print("Inside Non of them") 
    print("INVOICE CATEGORY: ")
    print(invoice_category)
    print("REQUESTED PRICE")
    print(requested_price)
    #branchProduct = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
    form.price.data = requested_price
    #form.available_quantity.data =  100
    list1 = session.get('cart')
    list_of_items = []
    print(list1)
    if list1 == None:
      list_of_items = []
      return render_template('sell-branch-2-loan.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
    for key, value in list1.items():
      temp = [key,value]
      list_of_items.append(temp)
    form = SellLoan()
    form.price.data = requested_price
    print("FORM DATA PASSES")
    print(form.price.data)
    #form.available_quantity = int(branchProduct.quantity)
    form.autocomp.data = branchProduct.name
    form.autocompcustomer.data = customer_name
    form.inv_category.data = invoice_category_arabic
    return render_template('sell-branch-2-loan.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
    #return invoice_category + str(requested_price)
  elif form.confirm.data:
    list1 = session.get('cart')
    list_of_items = []
    for key, value in list1.items():
      temp = [key,value]
      list_of_items.append(temp)
      branchProduct = BranchTwoProduct.query.filter_by(name=key).first()
      branchProduct.quantity = branchProduct.quantity - int(value[1])
      db.session.commit()
    y = json.dumps(list1)
    print(y)
    total = 0
    for item in list_of_items:
      total = total + (item[1][0] * item[1][1])
    print("The total of invoice nad transactions is ...")
    print(total)
    #(percent * whole) / 100.0
    #vat_percentage = VAT(vat = 15)
    #db.session.add(vat_percentage)
    #db.session.commit()
    vat_value = VAT.query.all()
    #print(vat_value)
    #print(vat_value[0])
    #print(vat_value[0].vat)
    vat_value = vat_value[0].vat
    print("&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&")
    #print("VAT NOT SUB")
    #print(vat_value[0])
    vat = vat_value * total / 100
    total = total + vat
    print("The list..")
    print(list_of_items)
    print("The total is ...")
    print(total)
    print(form.autocompcustomer.data)
    c = Customer.query.filter_by(name=form.autocompcustomer.data).first()
    percentage = VAT.query.filter_by(id=1).first()
    print(percentage)
    print(percentage.vat)
    now = datetime.now()
    inv = Inv(products=y, vat_value=vat, vat_percentage= percentage.vat,  total=total, inv_type = "Loan", status = "Not Paid", remaining_balance=total, initiator = "Branch2" ,category= request.form.get('inv_category'), date = now, customer_id = c.id)
    db.session.add(inv)
    db.session.commit()
    db.session.flush()
    print("LAST COMMITED INVOICE")
    print(inv.id)
    print(Inv.query.all())
    invvv= Inv.query.all()
    print(invvv[0].products)
    c.remaining_balance = c.remaining_balance + total
    db.session.commit()
    return redirect(url_for('invoice', invoice_id=inv.id))  
  else:
    list1 = session.get('cart')
    list_of_items = []
    print(list1)
    for key, value in list1.items():
      temp = [key,value]
      list_of_items.append(temp)   
    print("CART CONTENT")
    print(list_of_items)  
    return render_template('sell-branch-2-loan.html', form=form, products = list_of_items, length = len(list_of_items), user=u);    
    #return "Finished"

  list1 = session.get('cart')
  list_of_items = []
  print(list1)
  for key, value in list1.items():
    temp = [key,value]
    list_of_items.append(temp)   
  print("CART CONTENT")
  print(list_of_items)  
  return render_template('sell-branch-2-loan.html', form=form, products = list_of_items, length = len(list_of_items), user=u);    
  #return "Finished"       


@app.route("/sell-branch-1-loan-to-dbb.html", methods=['GET', 'POST'])
@login_required
def sellBranchOneLoanDB():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  list_of_items = []
  print("my func!!")
  form = SellLoan(request.form)
  print("Inside add to DB")
  print(form.errors)
  if form.validate():
    if form.view.data:
      sellingHistory("Wessam Gholam")
      return "View button is clicked"
    elif form.submit.data:
        print("Submit button is clicked")
        return "Submit button is clicked"
    elif form.get_price.data:
      print("Get Price button is clicked")
      branchProduct = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
      print(request.form.get('inv_category'))
      return "sure"
      #form.inv_category.data
      #form.price.data = 
    #elif form.submit.data:
      #print("Submit button is clicked")
    elif form.clear.data:  
      print("Clear button is clicked")
      session.pop('cart')
      #session.pop('cart', None)
      #session['cart'] = {}
      list1 = session.get('cart')
      list_of_items = []
      print(list1)
      if list1 == None:
        list_of_items = []
        return render_template('sell-branch-1-loan.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
      for key, value in list1.items():
        temp = [key,value]
        list_of_items.append(temp)
      return render_template('sell-branch-1-loan.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
    elif form.confirm.data:
      list1 = session.get('cart')
      list_of_items = []
      for key, value in list1.items():
        temp = [key,value]
        list_of_items.append(temp)
      y = json.dumps(list1)
      print(y)
      total = 0
      for item in list_of_items:
        total = total + (item[1][0] * item[1][1])
      print("The total of invoice nad transactions is ...")
      print(total)
      #(percent * whole) / 100.0
      #vat_percentage = VAT(vat = 15)
      #db.session.add(vat_percentage)
      #db.session.commit()
      vat_value = VAT.query.all()
      print(vat_value)
      vat_value = vat_value[0].vat
      vat = vat_value * total / 100
      total = total + vat
      print("The list..")
      print(list_of_items)
      print("The total is ...")
      print(total)
      print(form.autocompcustomer.data)
      c = Customer.query.filter_by(name=form.autocompcustomer.data).first()
      inv = Inv(products=y, vat_value=vat, total=total, inv_type = "Loan", status = "Not Paid", remaining_balance=total, customer_id = c.id)
      db.session.add(inv)
      db.session.commit()
      db.session.flush()
      print("LAST COMMITED INVOICE")
      print(inv.id)
      print(Inv.query.all())
      invvv= Inv.query.all()
      print(invvv[0].products)
      return redirect(url_for('invoice', invoice_id=inv.id))
      #return render_template('invoice.html',products=list_of_items, length = len(list_of_items))
    elif form.view.data:
      return "View button is clicked"
    elif form.submit.data:
      print("Submit button is clicked")      
      print("inside form validation")
      print(form.autocomp.data)
      print(form.quantity.data)
      print(form.price.data)
      pb1 = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
      p = Product.query.filter_by(name=form.autocomp.data).first()
      print("Quantity check ...")
      print(form.quantity.data)
      print(type(form.quantity.data))
      #print(type(p.b1_quantity))
      #print(p.b1_quantity)
      if form.quantity.data >= pb1.quantity:
        flash(u'Product Quantity is out of stock', 'danger')
        return redirect('sell-branch-1-loan.html')
      #p = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
      p_all = BranchOneProduct.query.all()
      print("all products")
      print(p_all)
      print("product")
      ##print(p)
      print("product price: ")
      ##print(p.price)
      print("product availability: ")
      ##print(p.quantity)
      ##print(type(p.quantity))
      print("product quantity")
      print(form.quantity.data)
      print(type(form.quantity.data))
      invoice = Invoice()
      db.session.add(invoice)
      db.session.commit()

      print("Invo query: ")
      print(Invoice.query.all())
      print("First Invoice is >>>")
      invoice = Invoice.query.filter_by(invoice_id=1).first()
      print(invoice)
      print("PRINT ALL PRODUCTS ASSOCIATED WITH THIS INVOICE:")
      print(invoice.products)
      print("THE END!")

      #p = Product(name="switches", price=10,shelf="W100", quantity=1000)

      db.session.add(p)
      db.session.commit()

      invoice.products.append(p)
      db.session.commit()

      print("Equiry joint table for INVOICES")
      ##print(Class.query.join(Class.students).all())
      print(invoice.products)

      now = datetime.now()
      #t = Transaction(t_type="CR", total=1000,date=now, description="buy stuff", balance=45000, p_type="Cash")
      #t.invoice = invoice 

      #db.session.add(t)
      #db.session.commit()

      #print("Transactions : ")
      #print(Transaction.query.filter_by(id=1).first())
      #print(Transaction.query.filter_by(id=1).first().invoice)

      """
      cart_item = {'pineapples': '10', 'apples': '20', 'mangoes': '30'}
      print("HERE I'M ...")
      print(cart_item)
      print(type(cart_item))
      items_list = []
      items_list.append(p)
      list_of_items=session.get('cart')
      session['cart'] = cart_item
      list_of_items=session.get('cart')
      print("DICT VALUES ARE: ....")
      print(session.get('cart'))
      print(len(list_of_items))
      """
      if 'cart' not in session:
        print("$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$")
        print("Creating session")
        session['cart'] = {}

      #session['cart'] = {}
      item = [ float(form.price.data), form.quantity.data]

      if item:
        cart_list = session['cart']
        print("CART: ")
        cart_list[form.autocomp.data] = item
        print(cart_list)
        print(type(cart_list))
        session['cart'] = cart_list
        list1 = session.get('cart')
        list_of_items = []
        print(list1)
        for key, value in list1.items():
          temp = [key,value]
          list_of_items.append(temp)

        print("The Full List: ")
        print(list_of_items)
        print(list_of_items[0][0])
        print(list_of_items[0][1])
        print(list_of_items[0][1][0])
        print(list_of_items[0][1][1])
        print(len(list_of_items))
        #str1 = ''.join(list_of_items)
        print("MY STRING IS")
      #product = Product(name=form.name.data, price=form.price.data,shelf=form.shelf.data, w_quantity=form.quantity.data, b1_quantity=0, b2_quantity=0)
      

      ##y = json.dumps(list1)
      ##print(y)
      ##inv = Inv(products=y)
      ##db.session.add(inv)
      ##db.session.commit()
      ##print(Inv.query.all())
      ##invvv= Inv.query.all()
      ##print(invvv[0].products)



      #products = db.Column(db.varchar(250))
      #db.session.add(product)
      #y = json.dumps(list1)
      #print(y)
      #print(request.form['clear'])
      #if request.form['submit'] is not None:
       # session.pop('cart')
        #print("Cart is cleared")




      #cart_list.append(item)
      #session['cart'] = cart_list  # 
      #session.modified = True
      #print("Shopping cart: ")
      #print(session.get('cart'))
      #print(session['cart'])
      #session['cart'] = items_list
      #p = jsonify(items_list)

      return render_template('sell-branch-1-loan.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  return render_template('sell-branch-1-loan.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
















@app.route("/sell-branch-1-cash-to-db.html", methods=['GET', 'POST'])
@login_required
def sellBranchOneCashDB():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  print("my func!!")
  form = SellCash(request.form)
  print("Inside add to DB")
  print(form.errors)
  invoice_category_arabic = ""
  customer_name = form.autocompcustomer.data
  #if form.validate():
  if form.submit.data:
    #return"test!!"
    print("Submit button is clicked")
    print("inside form validation")
    print(form.autocomp.data)
    print(form.quantity.data)
    print(form.price.data)
    p = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
    print("Quantity check ...")
    print(form.quantity.data)
    print(type(form.quantity.data))
    print(type(p.quantity))
    print(p.quantity)
    list1 = session.get('cart')
    print("CHECK CART SESSION")
    print(list1)
    list_of_items = []
    if form.quantity.data >= p.quantity:
      flash(u'Product Quantity is out of stock', 'danger')
      list1 = session.get('cart')
      list_of_items = []
      print(list1)
      if list1 == None:
        list_of_items = []
        return render_template('sell-branch-1-cash.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
    if 'cart' not in session:
      print("$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$")
      print("Creating session")
      session['cart'] = {}
    #session['cart'] = {}
    item = [ float(form.price.data), form.quantity.data]
    if item:
      cart_list = session['cart']
      print("CART: ")
      cart_list[form.autocomp.data] = item
      print(cart_list)
      print(type(cart_list))
    session['cart'] = cart_list
    list1 = session.get('cart')
    list_of_items = []
    print(list1)
    for key, value in list1.items():
      temp = [key,value]
      list_of_items.append(temp)   
    print("CART CONTENT")
    print(list_of_items)  
    return render_template('sell-branch-1-cash.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
      #return redirect('sell-branch-1-cash.html')
    #p = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
    p_all = BranchOneProduct.query.all()
    print("all products")
    print(p_all)
    print("product")
    ##print(p)
    print("product price: ")
    ##print(p.price)
    print("product availability: ")
    ##print(p.quantity)
    ##print(type(p.quantity))
    print("product quantity")
    print(form.quantity.data)
    print(type(form.quantity.data))
    invoice = Invoice()
    db.session.add(invoice)
    db.session.commit()

    print("Invo query: ")
    print(Invoice.query.all())
    print("First Invoice is >>>")
    invoice = Invoice.query.filter_by(invoice_id=1).first()
    print(invoice)
    print("PRINT ALL PRODUCTS ASSOCIATED WITH THIS INVOICE:")
    print(invoice.products)
    print("THE END!")

    db.session.add(p)
    db.session.commit()

    invoice.products.append(p)
    db.session.commit()

    print("Equiry joint table for INVOICES")
    ##print(Class.query.join(Class.students).all())
    print(invoice.products)

    now = datetime.now()
    t = Transaction(t_type="CR", total=1000,date=now, description="buy stuff", balance=45000, p_type="Cash")
    t.invoice = invoice 

    db.session.add(t)
    db.session.commit()

    print("Transactions : ")
    print(Transaction.query.filter_by(id=1).first())
    print(Transaction.query.filter_by(id=1).first().invoice)
    return render_template('sell-branch-1-cash.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  elif form.clear.data:  
    print("Clear button is clicked")
    session.pop('cart')
    #session.pop('cart', None)
    #session['cart'] = {}
    list1 = session.get('cart')
    list_of_items = []
    print(list1)
    if list1 == None:
      list_of_items = []
      return render_template('sell-branch-1-cash.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
    for key, value in list1.items():
      temp = [key,value]
      list_of_items.append(temp)
    return render_template('sell-branch-1-cash.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  elif form.get_price.data:
    print("Get Price button is clicked")
    #product_name = 
    branchProduct = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
    print(request.form.get('inv_category'))
    invoice_category = ""
    invoice_category_arabic = ""
    requested_price = 0
    if request.form.get('inv_category') == "جملة " or request.form.get('inv_category')  == "Bulk":
      print("Inside bulk")
      invoice_category = "bulk_price"
      invoice_category_arabic = "جملة "
      requested_price = branchProduct.bulk_price
    elif request.form.get('inv_category') == "جملة  الجملة " or request.form.get('inv_category')  == "Bulk Bulk":
      print("Inside bulk bulk")
      invoice_category = "bulk_bulk_price"
      invoice_category_arabic = "جملة  الجملة "
      requested_price = branchProduct.bulk_bulk_price
    elif request.form.get('inv_category') == "تجزئة " or request.form.get('inv_category')  == "Single":
      print("Inside single")
      invoice_category = "single_price"
      invoice_category_arabic = "تجزئة "
      requested_price = branchProduct.single_price
    else:
      print("Inside Non of them")  

    print("INVOICE CATEGORY: ")
    print(invoice_category)
    print("REQUESTED PRICE")
    print(requested_price)
    #branchProduct = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
    form.price.data = requested_price
    #form.available_quantity.data =  100
    list1 = session.get('cart')
    list_of_items = []
    print(list1)
    if list1 == None:
      list_of_items = []
      return render_template('sell-branch-1-cash.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
    for key, value in list1.items():
      temp = [key,value]
      list_of_items.append(temp)
    form = SellLoan()
    form.price.data = requested_price
    print("FORM DATA PASSES")
    print(form.price.data)
    #form.available_quantity = int(branchProduct.quantity)
    form.autocomp.data = branchProduct.name
    form.autocompcustomer.data = customer_name
    form.inv_category.data = invoice_category_arabic
    return render_template('sell-branch-1-cash.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  elif form.confirm.data:
    list1 = session.get('cart')
    list_of_items = []
    for key, value in list1.items():
      temp = [key,value]
      list_of_items.append(temp)
      branchProduct = BranchOneProduct.query.filter_by(name=key).first()
      branchProduct.quantity = branchProduct.quantity - int(value[1])
      db.session.commit()
    y = json.dumps(list1)
    print(y)
    total = 0
    for item in list_of_items:
      total = total + (item[1][0] * item[1][1])
    print("The total of invoice nad transactions is ...")
    print(total)
    #(percent * whole) / 100.0
    #vat_percentage = VAT(vat = 15)
    #db.session.add(vat_percentage)
    #db.session.commit()
    vat_value = VAT.query.all()
    print(vat_value)
    vat_value = vat_value[0].vat
    vat = vat_value * total / 100
    total = total + vat
    print("The list..")
    print(list_of_items)
    print("The total is ...")
    print(total)
    print(form.autocompcustomer.data)
    c = Customer.query.filter_by(name=form.autocompcustomer.data).first()
    percentage = VAT.query.filter_by(id=1).first()
    print(percentage)
    print(percentage.vat)
    now = datetime.now()
    inv = Inv(products=y, vat_value=vat, vat_percentage= percentage.vat, total=total, inv_type = "Cash", status = "Paid", remaining_balance=total, initiator = "Branch1", category= request.form.get('inv_category'), date=now,  customer_id = c.id)
    db.session.add(inv)
    db.session.commit()
    db.session.flush()
    print("LAST COMMITED INVOICE")
    print(inv.id)
    print(Inv.query.all())
    invvv= Inv.query.all()
    print(invvv[0].products)
    return redirect(url_for('invoice', invoice_id=inv.id))
    #return render_template('invoice.html',products=list_of_items, length = len(list_of_items))

  list1 = session.get('cart')
  list_of_items = []
  print(list1)
  for key, value in list1.items():
    temp = [key,value]
    list_of_items.append(temp)   
  print("CART CONTENT")
  print(list_of_items)  
  return render_template('sell-branch-1-cash.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  #return "okie dokie"
  



@app.route("/sell-branch-2-cash-to-db.html", methods=['GET', 'POST'])
@login_required
def sellBranchTwoCashDB():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  print("my func!!")
  form = SellCash(request.form)
  print("Inside add to DB")
  print(form.errors)
  invoice_category_arabic = ""
  customer_name = form.autocompcustomer.data
  #if form.validate():
  if form.submit.data:
    #return"test!!"
    print("Submit button is clicked")
    print("inside form validation")
    print(form.autocomp.data)
    print(form.quantity.data)
    print(form.price.data)
    p = BranchTwoProduct.query.filter_by(name=form.autocomp.data).first()
    print("Quantity check ...")
    print(form.quantity.data)
    print(type(form.quantity.data))
    print(type(p.quantity))
    print(p.quantity)
    list1 = session.get('cart')
    print("CHECK CART SESSION")
    print(list1)
    list_of_items = []
    if form.quantity.data > p.quantity:
      flash(u'Product Quantity is out of stock', 'danger')
      list1 = session.get('cart')
      list_of_items = []
      print(list1)
      if list1 == None:
        list_of_items = []
        return render_template('sell-branch-2-cash.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
    if 'cart' not in session:
      print("$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$")
      print("Creating session")
      session['cart'] = {}
    #session['cart'] = {}
    item = [ float(form.price.data), form.quantity.data]
    if item:
      cart_list = session['cart']
      print("CART: ")
      cart_list[form.autocomp.data] = item
      print(cart_list)
      print(type(cart_list))
    session['cart'] = cart_list
    list1 = session.get('cart')
    list_of_items = []
    print(list1)
    for key, value in list1.items():
      temp = [key,value]
      list_of_items.append(temp)   
    print("CART CONTENT")
    print(list_of_items)  
    return render_template('sell-branch-2-cash.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
      #return redirect('sell-branch-1-cash.html')
    #p = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
    p_all = BranchTwoProduct.query.all()
    print("all products")
    print(p_all)
    print("product")
    ##print(p)
    print("product price: ")
    ##print(p.price)
    print("product availability: ")
    ##print(p.quantity)
    ##print(type(p.quantity))
    print("product quantity")
    print(form.quantity.data)
    print(type(form.quantity.data))
    invoice = Invoice()
    db.session.add(invoice)
    db.session.commit()

    print("Invo query: ")
    print(Invoice.query.all())
    print("First Invoice is >>>")
    invoice = Invoice.query.filter_by(invoice_id=1).first()
    print(invoice)
    print("PRINT ALL PRODUCTS ASSOCIATED WITH THIS INVOICE:")
    print(invoice.products)
    print("THE END!")

    db.session.add(p)
    db.session.commit()

    invoice.products.append(p)
    db.session.commit()

    print("Equiry joint table for INVOICES")
    ##print(Class.query.join(Class.students).all())
    print(invoice.products)

    now = datetime.now()
    t = Transaction(t_type="CR", total=1000,date=now, description="buy stuff", balance=45000, p_type="Cash")
    t.invoice = invoice 

    db.session.add(t)
    db.session.commit()

    print("Transactions : ")
    print(Transaction.query.filter_by(id=1).first())
    print(Transaction.query.filter_by(id=1).first().invoice)
    return render_template('sell-branch-2-cash.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  elif form.clear.data:  
    print("Clear button is clicked")
    session.pop('cart')
    #session.pop('cart', None)
    #session['cart'] = {}
    list1 = session.get('cart')
    list_of_items = []
    print(list1)
    if list1 == None:
      list_of_items = []
      return render_template('sell-branch-2-cash.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
    for key, value in list1.items():
      temp = [key,value]
      list_of_items.append(temp)
    return render_template('sell-branch-2-cash.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  elif form.get_price.data:
    print("Get Price button is clicked")
    #product_name = 
    branchProduct = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
    print(request.form.get('inv_category'))
    invoice_category = ""
    invoice_category_arabic = ""
    requested_price = 0
    if request.form.get('inv_category') == "جملة " or request.form.get('inv_category')  == "Bulk":
      print("Inside bulk")
      invoice_category = "bulk_price"
      invoice_category_arabic = "جملة "
      requested_price = branchProduct.bulk_price
    elif request.form.get('inv_category') == "جملة  الجملة " or request.form.get('inv_category')  == "Bulk Bulk":
      print("Inside bulk bulk")
      invoice_category = "bulk_bulk_price"
      invoice_category_arabic = "جملة  الجملة "
      requested_price = branchProduct.bulk_bulk_price
    elif request.form.get('inv_category') == "تجزئة " or request.form.get('inv_category')  == "Single":
      print("Inside single")
      invoice_category = "single_price"
      invoice_category_arabic = "تجزئة "
      requested_price = branchProduct.single_price
    else:
      print("Inside Non of them") 
    print("INVOICE CATEGORY: ")
    print(invoice_category)
    print("REQUESTED PRICE")
    print(requested_price)
    #branchProduct = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
    form.price.data = requested_price
    #form.available_quantity.data =  100
    list1 = session.get('cart')
    list_of_items = []
    print(list1)
    if list1 == None:
      list_of_items = []
      return render_template('sell-branch-2-cash.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
    for key, value in list1.items():
      temp = [key,value]
      list_of_items.append(temp)
    form = SellLoan()
    form.price.data = requested_price
    print("FORM DATA PASSES")
    print(form.price.data)
    #form.available_quantity = int(branchProduct.quantity)
    form.autocomp.data = branchProduct.name
    form.autocompcustomer.data = customer_name
    form.inv_category.data = invoice_category_arabic
    return render_template('sell-branch-2-cash.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  elif form.confirm.data:
    list1 = session.get('cart')
    list_of_items = []
    for key, value in list1.items():
      temp = [key,value]
      list_of_items.append(temp)
      branchProduct = BranchTwoProduct.query.filter_by(name=key).first()
      branchProduct.quantity = branchProduct.quantity - int(value[1])
      db.session.commit()
    y = json.dumps(list1)
    print(y)
    total = 0
    for item in list_of_items:
      total = total + (item[1][0] * item[1][1])
    print("The total of invoice nad transactions is ...")
    print(total)
    #(percent * whole) / 100.0
    #vat_percentage = VAT(vat = 15)
    #db.session.add(vat_percentage)
    #db.session.commit()
    vat_value = VAT.query.all()
    print(vat_value)
    vat_value = vat_value[0].vat
    vat = vat_value * total / 100
    total = total + vat
    print("The list..")
    print(list_of_items)
    print("The total is ...")
    print(total)
    print(form.autocompcustomer.data)
    c = Customer.query.filter_by(name=form.autocompcustomer.data).first()
    percentage = VAT.query.filter_by(id=1).first()
    print(percentage)
    print(percentage.vat)
    now = datetime.now()
    inv = Inv(products=y, vat_value=vat, vat_percentage= percentage.vat, total=total, inv_type = "Cash", status = "Paid", remaining_balance=None, initiator = "Branch2", category= request.form.get('inv_category'), date=now,  customer_id = c.id)
    db.session.add(inv)
    db.session.commit()
    db.session.flush()
    print("LAST COMMITED INVOICE")
    print(inv.id)
    print(Inv.query.all())
    invvv= Inv.query.all()
    print(invvv[0].products)
    return redirect(url_for('invoice', invoice_id=inv.id))
    #return render_template('invoice.html',products=list_of_items, length = len(list_of_items))

  list1 = session.get('cart')
  list_of_items = []
  print(list1)
  for key, value in list1.items():
    temp = [key,value]
    list_of_items.append(temp)   
  print("CART CONTENT")
  print(list_of_items)  
  return render_template('sell-branch-2-cash.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  #return "okie dokie"
  









@app.route("/sell-branch-2-card-to-db.html", methods=['GET', 'POST'])
@login_required
def sellBranchTwoCardDB():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  print("my func!!")
  form = SellCash(request.form)
  print("Inside add to DB")
  print(form.errors)
  invoice_category_arabic = ""
  customer_name = form.autocompcustomer.data
  #if form.validate():
  if form.submit.data:
    #return"test!!"
    print("Submit button is clicked")
    print("inside form validation")
    print(form.autocomp.data)
    print(form.quantity.data)
    print(form.price.data)
    p = BranchTwoProduct.query.filter_by(name=form.autocomp.data).first()
    print("Quantity check ...")
    print(form.quantity.data)
    print(type(form.quantity.data))
    print(type(p.quantity))
    print(p.quantity)
    list1 = session.get('cart')
    print("CHECK CART SESSION")
    print(list1)
    list_of_items = []
    if form.quantity.data >= p.quantity:
      flash(u'Product Quantity is out of stock', 'danger')
      list1 = session.get('cart')
      list_of_items = []
      print(list1)
      if list1 == None:
        list_of_items = []
        return render_template('sell-branch-2-card.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
    if 'cart' not in session:
      print("$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$")
      print("Creating session")
      session['cart'] = {}
    #session['cart'] = {}
    item = [ float(form.price.data), form.quantity.data]
    if item:
      cart_list = session['cart']
      print("CART: ")
      cart_list[form.autocomp.data] = item
      print(cart_list)
      print(type(cart_list))
    session['cart'] = cart_list
    list1 = session.get('cart')
    list_of_items = []
    print(list1)
    for key, value in list1.items():
      temp = [key,value]
      list_of_items.append(temp)   
    print("CART CONTENT")
    print(list_of_items)  
    return render_template('sell-branch-2-card.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
      #return redirect('sell-branch-1-cash.html')
    #p = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
    p_all = BranchTwoProduct.query.all()
    print("all products")
    print(p_all)
    print("product")
    ##print(p)
    print("product price: ")
    ##print(p.price)
    print("product availability: ")
    ##print(p.quantity)
    ##print(type(p.quantity))
    print("product quantity")
    print(form.quantity.data)
    print(type(form.quantity.data))
    invoice = Invoice()
    db.session.add(invoice)
    db.session.commit()

    print("Invo query: ")
    print(Invoice.query.all())
    print("First Invoice is >>>")
    invoice = Invoice.query.filter_by(invoice_id=1).first()
    print(invoice)
    print("PRINT ALL PRODUCTS ASSOCIATED WITH THIS INVOICE:")
    print(invoice.products)
    print("THE END!")

    db.session.add(p)
    db.session.commit()

    invoice.products.append(p)
    db.session.commit()

    print("Equiry joint table for INVOICES")
    ##print(Class.query.join(Class.students).all())
    print(invoice.products)

    now = datetime.now()
    t = Transaction(t_type="CR", total=1000,date=now, description="buy stuff", balance=45000, p_type="Cash")
    t.invoice = invoice 

    db.session.add(t)
    db.session.commit()

    print("Transactions : ")
    print(Transaction.query.filter_by(id=1).first())
    print(Transaction.query.filter_by(id=1).first().invoice)
    return render_template('sell-branch-2-card.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  elif form.clear.data:  
    print("Clear button is clicked")
    session.pop('cart')
    #session.pop('cart', None)
    #session['cart'] = {}
    list1 = session.get('cart')
    list_of_items = []
    print(list1)
    if list1 == None:
      list_of_items = []
      return render_template('sell-branch-2-card.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
    for key, value in list1.items():
      temp = [key,value]
      list_of_items.append(temp)
    return render_template('sell-branch-2-card.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  elif form.get_price.data:
    print("Get Price button is clicked")
    #product_name = 
    branchProduct = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
    print(request.form.get('inv_category'))
    invoice_category = ""
    invoice_category_arabic = ""
    requested_price = 0
    if request.form.get('inv_category') == "جملة " or request.form.get('inv_category')  == "Bulk":
      print("Inside bulk")
      invoice_category = "bulk_price"
      invoice_category_arabic = "جملة "
      requested_price = branchProduct.bulk_price
    elif request.form.get('inv_category') == "جملة  الجملة " or request.form.get('inv_category')  == "Bulk Bulk":
      print("Inside bulk bulk")
      invoice_category = "bulk_bulk_price"
      invoice_category_arabic = "جملة  الجملة "
      requested_price = branchProduct.bulk_bulk_price
    elif request.form.get('inv_category') == "تجزئة " or request.form.get('inv_category')  == "Single":
      print("Inside single")
      invoice_category = "single_price"
      invoice_category_arabic = "تجزئة "
      requested_price = branchProduct.single_price
    else:
      print("Inside Non of them") 
    print("INVOICE CATEGORY: ")
    print(invoice_category)
    print("REQUESTED PRICE")
    print(requested_price)
    #branchProduct = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
    form.price.data = requested_price
    #form.available_quantity.data =  100
    list1 = session.get('cart')
    list_of_items = []
    print(list1)
    if list1 == None:
      list_of_items = []
      return render_template('sell-branch-2-card.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
    for key, value in list1.items():
      temp = [key,value]
      list_of_items.append(temp)
    form = SellLoan()
    form.price.data = requested_price
    print("FORM DATA PASSES")
    print(form.price.data)
    #form.available_quantity = int(branchProduct.quantity)
    form.autocomp.data = branchProduct.name
    form.autocompcustomer.data = customer_name
    form.inv_category.data = invoice_category_arabic
    return render_template('sell-branch-2-card.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  elif form.confirm.data:
    list1 = session.get('cart')
    list_of_items = []
    for key, value in list1.items():
      temp = [key,value]
      list_of_items.append(temp)
      branchProduct = BranchTwoProduct.query.filter_by(name=key).first()
      branchProduct.quantity = branchProduct.quantity - int(value[1])
      db.session.commit()
    y = json.dumps(list1)
    print(y)
    total = 0
    for item in list_of_items:
      total = total + (item[1][0] * item[1][1])
    print("The total of invoice nad transactions is ...")
    print(total)
    #(percent * whole) / 100.0
    #vat_percentage = VAT(vat = 15)
    #db.session.add(vat_percentage)
    #db.session.commit()
    vat_value = VAT.query.all()
    print(vat_value)
    vat_value = vat_value[0].vat
    vat = vat_value * total / 100
    total = total + vat
    print("The list..")
    print(list_of_items)
    print("The total is ...")
    print(total)
    print(form.autocompcustomer.data)
    c = Customer.query.filter_by(name=form.autocompcustomer.data).first()
    percentage = VAT.query.filter_by(id=1).first()
    print(percentage)
    print(percentage.vat)
    now = datetime.now()
    inv = Inv(products=y, vat_value=vat, vat_percentage= percentage.vat, total=total, inv_type = "شبكة", status = "Paid", remaining_balance=None, initiator = "Branch1", category= request.form.get('inv_category'), date=now,  customer_id = c.id)
    db.session.add(inv)
    db.session.commit()
    db.session.flush()
    print("LAST COMMITED INVOICE")
    print(inv.id)
    print(Inv.query.all())
    invvv= Inv.query.all()
    print(invvv[0].products)
    return redirect(url_for('invoice', invoice_id=inv.id))
    #return render_template('invoice.html',products=list_of_items, length = len(list_of_items))

  list1 = session.get('cart')
  list_of_items = []
  print(list1)
  for key, value in list1.items():
    temp = [key,value]
    list_of_items.append(temp)   
  print("CART CONTENT")
  print(list_of_items)  
  return render_template('sell-branch-2-card.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  #return "okie dokie"
  





@app.route("/sell-branch-1-card-to-db.html", methods=['GET', 'POST'])
@login_required
def sellBranchOneCardDB():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  print("my func!!")
  form = SellCash(request.form)
  print("Inside add to DB")
  print(form.errors)
  invoice_category_arabic = ""
  customer_name = form.autocompcustomer.data
  #if form.validate():
  if form.submit.data:
    #return"test!!"
    print("Submit button is clicked")
    print("inside form validation")
    print(form.autocomp.data)
    print(form.quantity.data)
    print(form.price.data)
    p = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
    print("Quantity check ...")
    print(form.quantity.data)
    print(type(form.quantity.data))
    print(type(p.quantity))
    print(p.quantity)
    list1 = session.get('cart')
    print("CHECK CART SESSION")
    print(list1)
    list_of_items = []
    if form.quantity.data >= p.quantity:
      flash(u'Product Quantity is out of stock', 'danger')
      list1 = session.get('cart')
      list_of_items = []
      print(list1)
      if list1 == None:
        list_of_items = []
        return render_template('sell-branch-1-card.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
    if 'cart' not in session:
      print("$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$")
      print("Creating session")
      session['cart'] = {}
    #session['cart'] = {}
    item = [ float(form.price.data), form.quantity.data]
    if item:
      cart_list = session['cart']
      print("CART: ")
      cart_list[form.autocomp.data] = item
      print(cart_list)
      print(type(cart_list))
    session['cart'] = cart_list
    list1 = session.get('cart')
    list_of_items = []
    print(list1)
    for key, value in list1.items():
      temp = [key,value]
      list_of_items.append(temp)   
    print("CART CONTENT")
    print(list_of_items)  
    return render_template('sell-branch-1-card.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
      #return redirect('sell-branch-1-cash.html')
    #p = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
    p_all = BranchOneProduct.query.all()
    print("all products")
    print(p_all)
    print("product")
    ##print(p)
    print("product price: ")
    ##print(p.price)
    print("product availability: ")
    ##print(p.quantity)
    ##print(type(p.quantity))
    print("product quantity")
    print(form.quantity.data)
    print(type(form.quantity.data))
    invoice = Invoice()
    db.session.add(invoice)
    db.session.commit()

    print("Invo query: ")
    print(Invoice.query.all())
    print("First Invoice is >>>")
    invoice = Invoice.query.filter_by(invoice_id=1).first()
    print(invoice)
    print("PRINT ALL PRODUCTS ASSOCIATED WITH THIS INVOICE:")
    print(invoice.products)
    print("THE END!")

    db.session.add(p)
    db.session.commit()

    invoice.products.append(p)
    db.session.commit()

    print("Equiry joint table for INVOICES")
    ##print(Class.query.join(Class.students).all())
    print(invoice.products)

    now = datetime.now()
    t = Transaction(t_type="CR", total=1000,date=now, description="buy stuff", balance=45000, p_type="Cash")
    t.invoice = invoice 

    db.session.add(t)
    db.session.commit()

    print("Transactions : ")
    print(Transaction.query.filter_by(id=1).first())
    print(Transaction.query.filter_by(id=1).first().invoice)
    return render_template('sell-branch-1-card.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  elif form.clear.data:  
    print("Clear button is clicked")
    session.pop('cart')
    #session.pop('cart', None)
    #session['cart'] = {}
    list1 = session.get('cart')
    list_of_items = []
    print(list1)
    if list1 == None:
      list_of_items = []
      return render_template('sell-branch-1-card.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
    for key, value in list1.items():
      temp = [key,value]
      list_of_items.append(temp)
    return render_template('sell-branch-1-card.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  elif form.get_price.data:
    print("Get Price button is clicked")
    #product_name = 
    branchProduct = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
    print(request.form.get('inv_category'))
    invoice_category = ""
    invoice_category_arabic = ""
    requested_price = 0
    if request.form.get('inv_category') == "جملة " or request.form.get('inv_category')  == "Bulk":
      print("Inside bulk")
      invoice_category = "bulk_price"
      invoice_category_arabic = "جملة "
      requested_price = branchProduct.bulk_price
    elif request.form.get('inv_category') == "جملة  الجملة " or request.form.get('inv_category')  == "Bulk Bulk":
      print("Inside bulk bulk")
      invoice_category = "bulk_bulk_price"
      invoice_category_arabic = "جملة  الجملة "
      requested_price = branchProduct.bulk_bulk_price
    elif request.form.get('inv_category') == "تجزئة " or request.form.get('inv_category')  == "Single":
      print("Inside single")
      invoice_category = "single_price"
      invoice_category_arabic = "تجزئة "
      requested_price = branchProduct.single_price
    else:
      print("Inside Non of them") 
    print("INVOICE CATEGORY: ")
    print(invoice_category)
    print("REQUESTED PRICE")
    print(requested_price)
    #branchProduct = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
    form.price.data = requested_price
    #form.available_quantity.data =  100
    list1 = session.get('cart')
    list_of_items = []
    print(list1)
    if list1 == None:
      list_of_items = []
      return render_template('sell-branch-1-card.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
    for key, value in list1.items():
      temp = [key,value]
      list_of_items.append(temp)
    form = SellLoan()
    form.price.data = requested_price
    print("FORM DATA PASSES")
    print(form.price.data)
    #form.available_quantity = int(branchProduct.quantity)
    form.autocomp.data = branchProduct.name
    form.autocompcustomer.data = customer_name
    form.inv_category.data = invoice_category_arabic
    return render_template('sell-branch-1-card.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  elif form.confirm.data:
    list1 = session.get('cart')
    list_of_items = []
    for key, value in list1.items():
      temp = [key,value]
      list_of_items.append(temp)
      branchProduct = BranchOneProduct.query.filter_by(name=key).first()
      branchProduct.quantity = branchProduct.quantity - int(value[1])
      db.session.commit()
    y = json.dumps(list1)
    print(y)
    total = 0
    for item in list_of_items:
      total = total + (item[1][0] * item[1][1])
    print("The total of invoice nad transactions is ...")
    print(total)
    #(percent * whole) / 100.0
    #vat_percentage = VAT(vat = 15)
    #db.session.add(vat_percentage)
    #db.session.commit()
    vat_value = VAT.query.all()
    print(vat_value)
    vat_value = vat_value[0].vat
    vat = vat_value * total / 100
    total = total + vat
    print("The list..")
    print(list_of_items)
    print("The total is ...")
    print(total)
    print(form.autocompcustomer.data)
    c = Customer.query.filter_by(name=form.autocompcustomer.data).first()
    percentage = VAT.query.filter_by(id=1).first()
    print(percentage)
    print(percentage.vat)
    now = datetime.now()
    inv = Inv(products=y, vat_value=vat, vat_percentage= percentage.vat, total=total, inv_type = "شبكة", status = "Paid", initiator = "Branch1", category= request.form.get('inv_category'), date=now,  customer_id = c.id)
    db.session.add(inv)
    db.session.commit()
    db.session.flush()
    print("LAST COMMITED INVOICE")
    print(inv.id)
    print(Inv.query.all())
    invvv= Inv.query.all()
    print(invvv[0].products)
    return redirect(url_for('invoice', invoice_id=inv.id))
    #return render_template('invoice.html',products=list_of_items, length = len(list_of_items))

  list1 = session.get('cart')
  list_of_items = []
  print(list1)
  for key, value in list1.items():
    temp = [key,value]
    list_of_items.append(temp)   
  print("CART CONTENT")
  print(list_of_items)  
  return render_template('sell-branch-1-card.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  #return "okie dokie"
  
  






@app.route("/sell-branch-2-kabs-to-db.html", methods=['GET', 'POST'])
@login_required
def sellBranchTwoKabsDB():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  print("my func!!")
  form = SellCash(request.form)
  print("Inside add to DB")
  print(form.errors)
  invoice_category_arabic = ""
  customer_name = form.autocompcustomer.data
  #if form.validate():
  if form.submit.data:
    #return"test!!"
    print("Submit button is clicked")
    print("inside form validation")
    print(form.autocomp.data)
    print(form.quantity.data)
    print(form.price.data)
    p = BranchTwoProduct.query.filter_by(name=form.autocomp.data).first()
    print("Quantity check ...")
    print(form.quantity.data)
    print(type(form.quantity.data))
    print(type(p.quantity))
    print(p.quantity)
    list1 = session.get('cart')
    print("CHECK CART SESSION")
    print(list1)
    list_of_items = []
    if form.quantity.data >= p.quantity:
      flash(u'Product Quantity is out of stock', 'danger')
      list1 = session.get('cart')
      list_of_items = []
      print(list1)
      if list1 == None:
        list_of_items = []
        return render_template('sell-branch-2-kabs.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
    if 'cart' not in session:
      print("$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$")
      print("Creating session")
      session['cart'] = {}
    #session['cart'] = {}
    item = [ float(form.price.data), form.quantity.data]
    if item:
      cart_list = session['cart']
      print("CART: ")
      cart_list[form.autocomp.data] = item
      print(cart_list)
      print(type(cart_list))
    session['cart'] = cart_list
    list1 = session.get('cart')
    list_of_items = []
    print(list1)
    for key, value in list1.items():
      temp = [key,value]
      list_of_items.append(temp)   
    print("CART CONTENT")
    print(list_of_items)  
    return render_template('sell-branch-2-kabs.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
      #return redirect('sell-branch-1-cash.html')
    #p = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
    p_all = BranchTwoProduct.query.all()
    print("all products")
    print(p_all)
    print("product")
    ##print(p)
    print("product price: ")
    ##print(p.price)
    print("product availability: ")
    ##print(p.quantity)
    ##print(type(p.quantity))
    print("product quantity")
    print(form.quantity.data)
    print(type(form.quantity.data))
    invoice = Invoice()
    db.session.add(invoice)
    db.session.commit()

    print("Invo query: ")
    print(Invoice.query.all())
    print("First Invoice is >>>")
    invoice = Invoice.query.filter_by(invoice_id=1).first()
    print(invoice)
    print("PRINT ALL PRODUCTS ASSOCIATED WITH THIS INVOICE:")
    print(invoice.products)
    print("THE END!")

    db.session.add(p)
    db.session.commit()

    invoice.products.append(p)
    db.session.commit()

    print("Equiry joint table for INVOICES")
    ##print(Class.query.join(Class.students).all())
    print(invoice.products)

    now = datetime.now()
    t = Transaction(t_type="CR", total=1000,date=now, description="buy stuff", balance=45000, p_type="Cash")
    t.invoice = invoice 

    db.session.add(t)
    db.session.commit()

    print("Transactions : ")
    print(Transaction.query.filter_by(id=1).first())
    print(Transaction.query.filter_by(id=1).first().invoice)
    return render_template('sell-branch-1-kabs.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  elif form.clear.data:  
    print("Clear button is clicked")
    session.pop('cart')
    #session.pop('cart', None)
    #session['cart'] = {}
    list1 = session.get('cart')
    list_of_items = []
    print(list1)
    if list1 == None:
      list_of_items = []
      return render_template('sell-branch-2-kabs.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
    for key, value in list1.items():
      temp = [key,value]
      list_of_items.append(temp)
    return render_template('sell-branch-2-kabs.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  elif form.get_price.data:
    print("Get Price button is clicked")
    #product_name = 
    branchProduct = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
    print(request.form.get('inv_category'))
    invoice_category = ""
    invoice_category_arabic = ""
    requested_price = 0
    if request.form.get('inv_category') == "جملة " or request.form.get('inv_category')  == "Bulk":
      print("Inside bulk")
      invoice_category = "bulk_price"
      invoice_category_arabic = "جملة "
      requested_price = branchProduct.bulk_price
    elif request.form.get('inv_category') == "جملة  الجملة " or request.form.get('inv_category')  == "Bulk Bulk":
      print("Inside bulk bulk")
      invoice_category = "bulk_bulk_price"
      invoice_category_arabic = "جملة  الجملة "
      requested_price = branchProduct.bulk_bulk_price
    elif request.form.get('inv_category') == "تجزئة " or request.form.get('inv_category')  == "Single":
      print("Inside single")
      invoice_category = "single_price"
      invoice_category_arabic = "تجزئة "
      requested_price = branchProduct.single_price
    else:
      print("Inside Non of them") 
    print("INVOICE CATEGORY: ")
    print(invoice_category)
    print("REQUESTED PRICE")
    print(requested_price)
    #branchProduct = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
    form.price.data = requested_price
    #form.available_quantity.data =  100
    list1 = session.get('cart')
    list_of_items = []
    print(list1)
    if list1 == None:
      list_of_items = []
      return render_template('sell-branch-2-kabs.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
    for key, value in list1.items():
      temp = [key,value]
      list_of_items.append(temp)
    form = SellLoan()
    form.price.data = requested_price
    print("FORM DATA PASSES")
    print(form.price.data)
    #form.available_quantity = int(branchProduct.quantity)
    form.autocomp.data = branchProduct.name
    form.autocompcustomer.data = customer_name
    form.inv_category.data = invoice_category_arabic
    return render_template('sell-branch-2-kabs.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  elif form.confirm.data:
    list1 = session.get('cart')
    list_of_items = []
    for key, value in list1.items():
      temp = [key,value]
      list_of_items.append(temp)
      branchProduct = BranchTwoProduct.query.filter_by(name=key).first()
      branchProduct.quantity = branchProduct.quantity - int(value[1])
      db.session.commit()
    y = json.dumps(list1)
    print(y)
    total = 0
    for item in list_of_items:
      total = total + (item[1][0] * item[1][1])
    print("The total of invoice nad transactions is ...")
    print(total)
    #(percent * whole) / 100.0
    #vat_percentage = VAT(vat = 15)
    #db.session.add(vat_percentage)
    #db.session.commit()
    vat_value = VAT.query.all()
    print(vat_value)
    vat_value = vat_value[0].vat
    vat = vat_value * total / 100
    total = total + vat
    print("The list..")
    print(list_of_items)
    print("The total is ...")
    print(total)
    print(form.autocompcustomer.data)
    c = Customer.query.filter_by(name=form.autocompcustomer.data).first()
    percentage = VAT.query.filter_by(id=1).first()
    print(percentage)
    print(percentage.vat)
    now = datetime.now()
    inv = Inv(products=y, vat_value=vat, vat_percentage= percentage.vat, total=total, inv_type = "كبس", status = "Paid", initiator = "Branch1", category= request.form.get('inv_category'), date=now,  customer_id = c.id)
    db.session.add(inv)
    db.session.commit()
    db.session.flush()
    print("LAST COMMITED INVOICE")
    print(inv.id)
    print(Inv.query.all())
    invvv= Inv.query.all()
    print(invvv[0].products)
    return redirect(url_for('invoice', invoice_id=inv.id))
    #return render_template('invoice.html',products=list_of_items, length = len(list_of_items))

  list1 = session.get('cart')
  list_of_items = []
  print(list1)
  for key, value in list1.items():
    temp = [key,value]
    list_of_items.append(temp)   
  print("CART CONTENT")
  print(list_of_items)  
  return render_template('sell-branch-2-kabs.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  #return "okie dokie"
  








@app.route("/sell-branch-1-kabs-to-db.html", methods=['GET', 'POST'])
@login_required
def sellBranchOneKabsDB():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  print("my func!!")
  form = SellCash(request.form)
  print("Inside add to DB")
  print(form.errors)
  invoice_category_arabic = ""
  customer_name = form.autocompcustomer.data
  #if form.validate():
  if form.submit.data:
    #return"test!!"
    print("Submit button is clicked")
    print("inside form validation")
    print(form.autocomp.data)
    print(form.quantity.data)
    print(form.price.data)
    p = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
    print("Quantity check ...")
    print(form.quantity.data)
    print(type(form.quantity.data))
    print(type(p.quantity))
    print(p.quantity)
    list1 = session.get('cart')
    print("CHECK CART SESSION")
    print(list1)
    list_of_items = []
    if form.quantity.data >= p.quantity:
      flash(u'Product Quantity is out of stock', 'danger')
      list1 = session.get('cart')
      list_of_items = []
      print(list1)
      if list1 == None:
        list_of_items = []
        return render_template('sell-branch-1-kabs.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
    if 'cart' not in session:
      print("$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$")
      print("Creating session")
      session['cart'] = {}
    #session['cart'] = {}
    item = [ float(form.price.data), form.quantity.data]
    if item:
      cart_list = session['cart']
      print("CART: ")
      cart_list[form.autocomp.data] = item
      print(cart_list)
      print(type(cart_list))
    session['cart'] = cart_list
    list1 = session.get('cart')
    list_of_items = []
    print(list1)
    for key, value in list1.items():
      temp = [key,value]
      list_of_items.append(temp)   
    print("CART CONTENT")
    print(list_of_items)  
    return render_template('sell-branch-1-kabs.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
      #return redirect('sell-branch-1-cash.html')
    #p = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
    p_all = BranchOneProduct.query.all()
    print("all products")
    print(p_all)
    print("product")
    ##print(p)
    print("product price: ")
    ##print(p.price)
    print("product availability: ")
    ##print(p.quantity)
    ##print(type(p.quantity))
    print("product quantity")
    print(form.quantity.data)
    print(type(form.quantity.data))
    invoice = Invoice()
    db.session.add(invoice)
    db.session.commit()

    print("Invo query: ")
    print(Invoice.query.all())
    print("First Invoice is >>>")
    invoice = Invoice.query.filter_by(invoice_id=1).first()
    print(invoice)
    print("PRINT ALL PRODUCTS ASSOCIATED WITH THIS INVOICE:")
    print(invoice.products)
    print("THE END!")

    db.session.add(p)
    db.session.commit()

    invoice.products.append(p)
    db.session.commit()

    print("Equiry joint table for INVOICES")
    ##print(Class.query.join(Class.students).all())
    print(invoice.products)

    now = datetime.now()
    t = Transaction(t_type="CR", total=1000,date=now, description="buy stuff", balance=45000, p_type="Cash")
    t.invoice = invoice 

    db.session.add(t)
    db.session.commit()

    print("Transactions : ")
    print(Transaction.query.filter_by(id=1).first())
    print(Transaction.query.filter_by(id=1).first().invoice)
    return render_template('sell-branch-1-kabs.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  elif form.clear.data:  
    print("Clear button is clicked")
    session.pop('cart')
    #session.pop('cart', None)
    #session['cart'] = {}
    list1 = session.get('cart')
    list_of_items = []
    print(list1)
    if list1 == None:
      list_of_items = []
      return render_template('sell-branch-1-kabs.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
    for key, value in list1.items():
      temp = [key,value]
      list_of_items.append(temp)
    return render_template('sell-branch-1-kabs.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  elif form.get_price.data:
    print("Get Price button is clicked")
    #product_name = 
    branchProduct = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
    print(request.form.get('inv_category'))
    invoice_category = ""
    invoice_category_arabic = ""
    requested_price = 0
    if request.form.get('inv_category') == "جملة " or request.form.get('inv_category')  == "Bulk":
      print("Inside bulk")
      invoice_category = "bulk_price"
      invoice_category_arabic = "جملة "
      requested_price = branchProduct.bulk_price
    elif request.form.get('inv_category') == "جملة  الجملة " or request.form.get('inv_category')  == "Bulk Bulk":
      print("Inside bulk bulk")
      invoice_category = "bulk_bulk_price"
      invoice_category_arabic = "جملة  الجملة "
      requested_price = branchProduct.bulk_bulk_price
    elif request.form.get('inv_category') == "تجزئة " or request.form.get('inv_category')  == "Single":
      print("Inside single")
      invoice_category = "single_price"
      invoice_category_arabic = "تجزئة "
      requested_price = branchProduct.single_price
    else:
      print("Inside Non of them") 
    print("INVOICE CATEGORY: ")
    print(invoice_category)
    print("REQUESTED PRICE")
    print(requested_price)
    #branchProduct = BranchOneProduct.query.filter_by(name=form.autocomp.data).first()
    form.price.data = requested_price
    #form.available_quantity.data =  100
    list1 = session.get('cart')
    list_of_items = []
    print(list1)
    if list1 == None:
      list_of_items = []
      return render_template('sell-branch-1-kabs.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
    for key, value in list1.items():
      temp = [key,value]
      list_of_items.append(temp)
    form = SellLoan()
    form.price.data = requested_price
    print("FORM DATA PASSES")
    print(form.price.data)
    #form.available_quantity = int(branchProduct.quantity)
    form.autocomp.data = branchProduct.name
    form.autocompcustomer.data = customer_name
    form.inv_category.data = invoice_category_arabic
    return render_template('sell-branch-1-kabs.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  elif form.confirm.data:
    list1 = session.get('cart')
    list_of_items = []
    for key, value in list1.items():
      temp = [key,value]
      list_of_items.append(temp)
      branchProduct = BranchOneProduct.query.filter_by(name=key).first()
      branchProduct.quantity = branchProduct.quantity - int(value[1])
      db.session.commit()
    y = json.dumps(list1)
    print(y)
    total = 0
    for item in list_of_items:
      total = total + (item[1][0] * item[1][1])
    print("The total of invoice nad transactions is ...")
    print(total)
    #(percent * whole) / 100.0
    #vat_percentage = VAT(vat = 15)
    #db.session.add(vat_percentage)
    #db.session.commit()
    vat_value = VAT.query.all()
    print(vat_value)
    vat_value = vat_value[0].vat
    vat = vat_value * total / 100
    total = total + vat
    print("The list..")
    print(list_of_items)
    print("The total is ...")
    print(total)
    print(form.autocompcustomer.data)
    c = Customer.query.filter_by(name=form.autocompcustomer.data).first()
    percentage = VAT.query.filter_by(id=1).first()
    print(percentage)
    print(percentage.vat)
    now = datetime.now()
    inv = Inv(products=y, vat_value=vat, vat_percentage= percentage.vat, total=total, inv_type = "كبس", status = "Paid", initiator = "Branch1", category= request.form.get('inv_category'), date=now,  customer_id = c.id)
    db.session.add(inv)
    db.session.commit()
    db.session.flush()
    print("LAST COMMITED INVOICE")
    print(inv.id)
    print(Inv.query.all())
    invvv= Inv.query.all()
    print(invvv[0].products)
    return redirect(url_for('invoice', invoice_id=inv.id))
    #return render_template('invoice.html',products=list_of_items, length = len(list_of_items))

  list1 = session.get('cart')
  list_of_items = []
  print(list1)
  for key, value in list1.items():
    temp = [key,value]
    list_of_items.append(temp)   
  print("CART CONTENT")
  print(list_of_items)  
  return render_template('sell-branch-1-kabs.html', form=form, products = list_of_items, length = len(list_of_items), user=u);
  #return "okie dokie"





@app.route("/edit-vat", methods=['GET', 'POST'])
@login_required
def editVAT():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  form = EditVAT(request.form)
  if form.submit.data:
    percentage = VAT.query.filter_by(id=1).first()
    percentage.vat = form.percentage.data
    db.session.commit()
    form.percentage.data = percentage.vat
    flash(u'تم تعديل قيمة الضريبة المضافة', 'success')
    return render_template('edit-vat.html', form=form, user=u)
  percentage = VAT.query.filter_by(id=1).first()
  form.percentage.data = percentage.vat 
  return render_template('edit-vat.html', form=form, user=u)

@app.route("/spendings", methods=['GET', 'POST'])
@login_required
def spendings():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  form = Spendings(request.form)
  if form.submit.data: 
    print(form.invoice_type.data)
    print(form.transaction_type.data)
    print(form.description.data)
    print(form.pay_amount.data)
    account = Account.query.filter_by(id=1).first()
    account.balance = account.balance - float(form.pay_amount.data)
    db.session.commit()
    print("Current Balance AFTER refund @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@: ")
    print(account.balance)
    now = datetime.now()
    my_user = current_user.get_id()
    print("Current user: ")
    print(my_user)
    initiator = ""
    u = User.query.filter_by(id=my_user).first()
    if u.admin == True:
      initiator = "admin"
    elif u.admin_alike:
      initiator = "admin_alike"
    elif u.warehouse:
      initiator = "warehouse"
    elif u.branch1:
      initiator = "branch1"
    elif u.branch2:
      initiator = "branch2"         
    inv = Inv(products="No", vat_value=0, vat_percentage= 0,  total=form.pay_amount.data, inv_type = "Cash", status = "Paid", initiator = initiator , category = "single" ,date = now, is_expense = True)
    #return "Sure"
    db.session.add(inv)
    db.session.commit()
    print("invoice id")
    print(inv.id)
    dr = DebitTransaction(t_type="DR", total=form.pay_amount.data, date=now, description=form.description.data , invoice_id=inv.id, current_balance = account.balance)
    db.session.add(dr)
    db.session.commit()
    flash(u'تم اضافة عملية الخصم', 'success')
    return render_template('spendings.html', form=form, user=u)
  return render_template('spendings.html', form=form, user=u)

@app.route("/view-invoices-admin", methods=['GET', 'POST'])
@login_required
def viewInvoicesAdmin():
  my_user = current_user.get_id()
  print("Inside View Invoices")
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  print(u.admin)
  invoices = Inv.query.join(Customer, Inv.customer_id==Customer.id).order_by(Inv.date).all()
  if u.admin:
    print("Admin and Branch1")
    invoices = Inv.query.join(Customer, Inv.customer_id==Customer.id).order_by(Inv.date).all()
    return render_template('view-invoices-admin.html', invoices=invoices, len=len(invoices), user=u) 
 

@app.route("/view-invoices", methods=['GET', 'POST'])
@login_required
def viewInvoices():
  my_user = current_user.get_id()
  print("Inside View Invoices")
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  print(u.admin)
  invoices = Inv.query.join(Customer, Inv.customer_id==Customer.id).order_by(Inv.date).all()
  if u.branch1 or u.admin:
    print("Admin and Branch1")
    invoices = Inv.query.filter_by(initiator="Branch1").join(Customer, Inv.customer_id==Customer.id).order_by(Inv.date).all()
  elif u.branch2 or u.admin:
     print("Admin or branch 2")
     invoices = Inv.query.filter_by(initiator="Branch2").join(Customer, Inv.customer_id==Customer.id).order_by(Inv.date).all() 
  return render_template('view-invoices.html', invoices=invoices, len=len(invoices), user=u) 

@app.route("/blank.html")
@login_required
def blank():
  return render_template('blank.html')


@app.route("/addToCart")
@login_required
def addToCart():
  return render_template('blank.html') 
  

@app.route("/add-products")
@login_required
def addProducts():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  print("Check ..")
  print(request.args)
  form = AddProductForm(request.form)
  return render_template('add-products.html', form=form, user=u)

@app.route("/add-customers")
@login_required
def addCustomers():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  print("Check ..")
  print(request.args)
  form = AddCustomerForm(request.form)
  return render_template('add-customers.html', form=form, user=u)  

@app.route("/add-customers-to-db", methods=['GET', 'POST'])
def addCustomersToDB():
  form = AddCustomerForm(request.form)
  print("Inside add to DB")
  print(form.errors)
  if form.validate():
    print("inside form validation")
    print(form.name.data)
    print(form.mobile.data)
    customer = Customer(name=form.name.data, mobile=form.mobile.data, remaining_balance = 0)
    db.session.add(customer)
    db.session.commit()
  return 'ok'

@app.route("/add-products-to-db", methods=['GET', 'POST'])
def addProductsToDB():
  form = AddProductForm(request.form)
  print("Inside add to DB")
  print(form.errors)
  if form.validate():
    print("inside form validation")
    print(form.name.data)
    print(form.shelf.data)
    print(form.quantity.data)
    print(":::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::")
    print(form.single_expense.data)
    print(form.bulk_bulk_expense.data)
    print(form.bulk_expense.data)
    #print(form.price.data)
    product = Product(name=form.name.data, bulk_price=form.bulk_price.data, bulk_bulk_price=form.bulk_bulk_price.data, single_price=form.single_price.data, single_expense= form.single_expense.data, bulk_bulk_expense = form.bulk_bulk_expense.data, bulk_expense = form.bulk_expense.data, shelf=form.shelf.data, quantity=form.quantity.data)
    productb1 = BranchOneProduct(name=form.name.data, bulk_price=form.bulk_price.data, bulk_bulk_price=form.bulk_bulk_price.data, single_price=form.single_price.data, single_expense= form.single_expense.data, bulk_bulk_expense = form.bulk_bulk_expense.data, bulk_expense = form.bulk_expense.data, shelf=form.shelf.data, quantity=0)
    productb2 = BranchTwoProduct(name=form.name.data, bulk_price=form.bulk_price.data, bulk_bulk_price=form.bulk_bulk_price.data, single_price=form.single_price.data, single_expense= form.single_expense.data, bulk_bulk_expense = form.bulk_bulk_expense.data, bulk_expense = form.bulk_expense.data, shelf=form.shelf.data, quantity=0)

    db.session.add(product)
    db.session.add(productb1)
    db.session.add(productb2)
    try:
      db.session.commit()
      flash(u'Product Added', 'success')
      p = Product.query.filter_by(product_id=1).first()
      print("Product:")
      print(p)
    except exc.SQLAlchemyError as e:
      error = str(e.__dict__['orig'])
      print(error)  
      flash(u'Product not added', 'danger')
  #return redirect('/add-products')
  return redirect(url_for('addProducts'))


@app.route("/amend-products-to-db", methods=['GET', 'POST'])
def amendProductsToDB():
  p = Product.query.filter_by(product_id=product_id).first()
  form = AmendProductForm(request.form)
  if form.submit.data:
    print("Submit button is clicked")
    product_id = form.id_number.data
    p.name = form.name.data
    p.bulk_price = form.bulk_price.data
    p.bulk_bulk_price = form.bulk_bulk_price.data
    p.single_price = form.single_price.data
    p.shelf = form.shelf.data 
    p.quantity = form.quantity.data
    db.session.commit()

    pb1 = BranchOneProduct.query.filter_by(name=form.name.data).first()
    print("Branch 1 Product")
    print(pb1)
    print(pb1.name)
    pb1.bulk_price = form.bulk_price.data
    pb1.bulk_bulk_price = form.bulk_bulk_price.data
    pb1.single_price = form.single_price.data
    db.session.commit()
    print("Price after commit: ")
    print(pb1.bulk_price)

    pb2 = BranchTwoProduct.query.filter_by(name=form.name.data).first()
    pb2.bulk_price = int(form.bulk_price.data)
    pb2.bulk_bulk_price = int(form.bulk_bulk_price.data)
    pb2.single_price = int(form.single_price.data)
    db.session.commit()

    p = Product.query.filter_by(product_id=product_id).first()
    form.id_number.data = product_id
    form.name.data = p.name
    form.bulk_price.data = p.bulk_price
    form.bulk_bulk_price.data = p.bulk_bulk_price
    form.single_price.data = p.single_price
    form.shelf.data = p.shelf
    form.quantity.data = p.quantity
    flash(u'تم تعديل معلومات الصنف', 'success')
    return redirect(url_for('amendProduct' ,product_id=p.product_id))
  #return "cool"     


@app.route("/register.html")
@login_required
def register():
  return render_template('register.html') 

@app.route("/buttons.html")
@login_required
def buttons():
  return render_template('buttons.html')

"""
@app.route('/<product_id>')
@login_required
def page(product_id):
    productid = product_id
    product = Product.query.filter_by(id=productid).first()
    form = AmendProductForm(request.form)
    form.id_number.data = product.id
    form.name.data = product.name
    form.price.data = product.price
    form.shelf.data = product.shelf
    form.quantity.data = product.quantity
    form.submit.label.text = "Amend Product"
    return render_template('amend.html', form=form)
    # You might want to return some sort of response...

"""

@app.route("/delete-product/<product_name>", methods=['GET', 'POST'])
@login_required
def deleteProduct(product_name):
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  print("The product name to be deleted is: ")
  print(product_name)
  list1 = session.get('cart')
  list1.pop(product_name)
  session['cart'] = list1
  print("::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::")
  redirect_to = request.referrer
  print("Previous Link: ")
  print(redirect_to)
  print(":>:>:>:>:>:>:>:>:>:>:>:>:>:>:>:>:>:>:>:>:>:>:>:>:>:>:>:>:>:>:>:>:>:>:>:>:>:>:>:>:>:>:>:>:>:>:>:")
  return redirect(redirect_to)
  #return "Page will be implemented"


@app.route("/amend-product/<product_id>", methods=['GET', 'POST'])
@login_required
def amendProduct(product_id):
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  #products = Product.query.all()
  print("AWWWWEHAAAA")
  p = Product.query.filter_by(product_id=product_id).first()
  pb1 = BranchOneProduct.query.filter_by(id=product_id).first()
  pb2 = BranchTwoProduct.query.filter_by(id=product_id).first()
  print(p)
  form = AmendProductForm(request.form)
  if form.submit.data:
    print("Submit button is clicked")
    print(form.bulk_price.data)
    print(form.name.data)
    product_id = form.id_number.data
    p.name = form.name.data
    pb1.name = form.name.data
    pb2.name = form.name.data
    p.bulk_price = form.bulk_price.data
    p.bulk_bulk_price = form.bulk_bulk_price.data
    p.single_price = form.single_price.data
    p.single_expense = form.single_expense.data
    p.bulk_bulk_expense = form.bulk_bulk_expense.data
    p.bulk_expense = form.bulk_expense.data
    p.shelf = form.shelf.data 
    p.quantity = form.quantity.data
    db.session.commit()
    #pb = Product.query.filter_by(product_id=product_id).first()

    pb1 = BranchOneProduct.query.filter_by(name=form.name.data).first()
    print("Branch 1 Product")
    print(pb1)
    print(pb1.name)
    pb1.bulk_price = form.bulk_price.data
    pb1.bulk_bulk_price = form.bulk_bulk_price.data
    pb1.single_price = form.single_price.data
    pb1.single_expense = form.single_expense.data
    pb1.bulk_bulk_expense = form.bulk_bulk_expense.data
    pb1.bulk_expense = form.bulk_expense.data
    db.session.commit()
    print("Price after commit: ")
    print(pb1.bulk_price)

    pb2 = BranchTwoProduct.query.filter_by(name=form.name.data).first()
    pb2.bulk_price = int(form.bulk_price.data)
    pb2.bulk_bulk_price = int(form.bulk_bulk_price.data)
    pb2.single_price = int(form.single_price.data)
    pb2.single_expense = form.single_expense.data
    pb2.bulk_bulk_expense = form.bulk_bulk_expense.data
    pb2.bulk_expense = form.bulk_expense.data
    db.session.commit()

    print(p.name)
    form.id_number.data = product_id
    form.name.data = p.name
    form.bulk_price.data = p.bulk_price
    form.bulk_bulk_price.data = p.bulk_bulk_price
    form.single_price.data = p.single_price
    form.shelf.data = p.shelf
    form.quantity.data = p.quantity
    form.bulk_expense.data = p.bulk_expense
    form.bulk_bulk_expense.data = p.bulk_bulk_expense
    form.single_expense.data = p.single_expense


    flash(u'تم تعديل معلومات الصنف', 'success')
    return render_template('amend-product.html', form=form, product_id= product_id, user=u)
  form.id_number.data = product_id
  form.name.data = p.name
  form.bulk_price.data = p.bulk_price
  form.bulk_bulk_price.data = p.bulk_bulk_price
  form.single_price.data = p.single_price
  form.shelf.data = p.shelf
  form.quantity.data = p.quantity
  form.bulk_expense.data = p.bulk_expense
  form.bulk_bulk_expense.data = p.bulk_bulk_expense
  form.single_expense.data = p.single_expense

  return render_template('amend-product.html', form=form, product_id= product_id, user=u)        



@app.route("/amend-products.html")
@login_required
def amednProducts():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  products = Product.query.all()
  return render_template('amend-products.html', products=products, len=len(products), user=u)        

@app.route("/tables.html")
@login_required
def tables():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  products = Product.query.all()
  print(products)
  print(len(products))
  return render_template('tables.html', products=products, len=len(products), user=u)

@app.route("/view-customers")
@login_required
def viewCustomers():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  cust = Customer.query.filter_by(id=1).first()
  print("Customer remaning balance: ")
  print(cust.remaining_balance)
  u = User.query.filter_by(id=my_user).first()
  customers = Customer.query.all()
  print(customers)
  print(len(customers))
  return render_template('view-customers.html', customers=customers, len=len(customers), user=u)    


def queryset_to_dict(query_result):
   query_columns = query_result[0].keys()
   res = [list(ele) for ele in query_result]
   dict_list = [dict(zip(query_columns, l)) for l in res]
   return dict_list


@app.route("/branch1products.html")
@login_required
def tablesB1():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  products = BranchOneProduct.query.all()
  #products2 = BranchOneProduct.query.join(Product, BranchOneProduct.name==Product.name).first()
  products2 = db.session.query(BranchOneProduct, Product).join(Product, (BranchOneProduct.name==Product.name)).all()

  print("After Join")
  print("::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::")
  #print(products2[0].Product.quantity)
  #invoices = Inv.query.filter_by(initiator="Branch1").join(Customer, Inv.customer_id==Customer.id).order_by(Inv.date).all()
  #print(products)
  print(len(products))
  return render_template('branch1products.html', products=products2, len=len(products2),user=u) 

@app.route("/branch2products.html")
@login_required
def tablesB2():
  my_user = current_user.get_id()
  print("Current user: ")
  print(my_user)
  u = User.query.filter_by(id=my_user).first()
  products = BranchTwoProduct.query.all()
  print("COMON LET SHCEK ")
  print(products)
  products2 = db.session.query(BranchTwoProduct, Product).join(Product, (BranchTwoProduct.name==Product.name)).all()
  print("check 2 ... ")
  print(products2)
  #print(len(products))
  return render_template('branch2products.html', products=products2, len=len(products2), user=u)     



#@LoginManager.unauthorized_handler     # In unauthorized_handler we have a callback URL 
#def unauthorized_callback():            # In call back url we can specify where we want to 
 # print("No access")
  #return redirect(url_for('login')) # redirect the user in my case it is login page!


def try_login(name,password):
    user = User.query.filter_by(username=name).first()
    #print(user.branch)
    print(password)
    print(user.hashed_password)
    print(sha256_crypt.verify(str(password), user.hashed_password))
    if user is not None:
      print("Not None")
      if user and sha256_crypt.verify(str(password), user.hashed_password):
        print("Passwords verified")
        login_user(user)
        return True
    return False

@lm.user_loader
def load_user(id):
  return User.query.get(int(id))